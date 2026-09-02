"""Temporal activity implementations for document-processing tasks.

The module name describes the domain work while the parent package makes the
Temporal SDK ownership explicit. Registered function names are intentionally
unchanged because they are part of persisted Temporal workflow history.
"""

import asyncio
import base64
import json
import logging
import os
import re
import shutil
import subprocess
import tempfile
import mimetypes
from dataclasses import replace
from datetime import datetime
from pathlib import Path
from typing import Callable
from uuid import uuid4

import httpx
import fitz
from langchain_text_splitters import RecursiveCharacterTextSplitter
from minio import Minio
from pypdf import PdfReader, PdfWriter
from temporalio import activity

from ..chunking import ChunkCandidate, ChunkingResult, chunk_pages, load_chunking_config
from ..chunking.enforce_limits import enforce_chunk_limits
from ..chunking.factory import is_llm_grouping_provider
from ..chunking.page_units import normalize_text
from ..ingestion_records import (
    _normalize_instance,
    clean_text,
    prepare_records as _prepare_records,
)
from ..ocr import ocr_pdf as run_ocr_pdf, ocr_pdf_in_segments as run_ocr_pdf_in_segments
from ..ocr.quality import degenerate_ocr_note, is_degenerate_repetition
from ..translation import load_translation_config, translate_pages
from ..vector_store import (
    get_vector_store,
    IndexSchemaReport,
    passage_index_settings,
    passage_schema_field_names,
    project_records,
)

SUPPORTED_INPUT_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".ppt", ".pptx", ".xls", ".xlsx", ".csv", ".jpg", ".jpeg", ".png", ".webp", ".tif", ".tiff"
}
IMAGE_INPUT_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".tif", ".tiff"}
OFFICE_INPUT_EXTENSIONS = {".doc", ".docx", ".ppt", ".pptx", ".xls", ".xlsx"}
DELIMITED_INPUT_EXTENSIONS = {".csv"}
NATIVE_SPREADSHEET_EXTENSIONS = {".xlsx"}

def get_minio_client():
    """Get MinIO client from environment. Credentials are required."""
    access_key = os.environ.get("MINIO_ACCESS_KEY")
    secret_key = os.environ.get("MINIO_SECRET_KEY")

    if not access_key or not secret_key:
        raise RuntimeError("MINIO_ACCESS_KEY and MINIO_SECRET_KEY environment variables are required")

    return Minio(
        os.environ.get("MINIO_ENDPOINT", "localhost:9000"),
        access_key=access_key,
        secret_key=secret_key,
        secure=False,
    )


def download_from_minio(minio_path: str) -> str:
    """Download file from MinIO and return local temp path."""
    path = minio_path.replace("minio://", "")
    parts = path.split("/", 1)
    bucket = parts[0]
    object_name = parts[1] if len(parts) > 1 else ""

    client = get_minio_client()

    suffix = Path(object_name).suffix
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    temp_path = temp_file.name
    temp_file.close()

    client.fget_object(bucket, object_name, temp_path)
    return temp_path


def _minio_object_name(instance: str | None, workflow_id: str, artifact_type: str, filename: str) -> str:
    """Object key for a document artifact, prefixed by its tenant for isolation.

    Layout: ``<instance>/<workflow_id>/<artifact_type>/<filename>``. New writes
    always carry the tenant prefix (including the default tenant). Reads never
    reconstruct this key — they use the stored ``minio://`` URI on the artifact
    row — so pre-prefix objects remain readable.
    """
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", filename or "artifact")
    inst = _normalize_instance(instance)
    return f"{inst}/{workflow_id}/{artifact_type}/{safe_name}"


def _upload_file_to_minio(
    local_path: str,
    workflow_id: str,
    artifact_type: str,
    filename: str,
    instance: str | None = None,
) -> tuple[str, int, str]:
    client = get_minio_client()
    bucket = os.environ.get("MINIO_BUCKET", "documents")
    object_name = _minio_object_name(instance, workflow_id, artifact_type, filename)
    mime_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    client.fput_object(bucket, object_name, local_path, content_type=mime_type)
    size_bytes = os.path.getsize(local_path)
    return (f"minio://{bucket}/{object_name}", size_bytes, mime_type)


def _write_json_temp(data: object, suffix: str = ".json") -> str:
    with tempfile.NamedTemporaryFile("w", delete=False, suffix=suffix, encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        return f.name


def _json_list(value: object) -> list:
    """Decode a stored JSON array column, tolerating null/blank/legacy values."""
    if isinstance(value, list):
        return value
    if not value:
        return []
    try:
        parsed = json.loads(value)
    except (TypeError, ValueError):
        return []
    return parsed if isinstance(parsed, list) else []


def _source_type_from_path(filepath: str) -> tuple[str, str]:
    ext = Path(filepath).suffix.lower()
    if ext in {".csv", ".xlsx"}:
        return ("spreadsheet", "spreadsheet")
    if ext in IMAGE_INPUT_EXTENSIONS | OFFICE_INPUT_EXTENSIONS | {".pdf"}:
        return ("document", "pdf")
    return ("unknown", "pdf")


def _normalized_filename(original_name: str, canonical_input_type: str) -> str:
    stem = Path(original_name).stem or "document"
    if canonical_input_type == "spreadsheet":
        ext = Path(original_name).suffix.lower()
        return f"{stem}{ext if ext in {'.csv', '.xlsx'} else '.csv'}"
    return f"{stem}.pdf"


def _normalized_artifact_type(canonical_input_type: str) -> str:
    if canonical_input_type == "spreadsheet":
        return "normalized_spreadsheet"
    return "normalized_pdf"


def _resolve_local_path(filepath: str) -> tuple[str, bool]:
    """
    Resolve path into local file path.
    Returns (path, should_cleanup).
    """
    if filepath.startswith("minio://") or filepath.startswith("minio:/"):
        minio_path = filepath
        if filepath.startswith("minio:/") and not filepath.startswith("minio://"):
            minio_path = filepath.replace("minio:/", "minio://", 1)
        local_path = download_from_minio(minio_path)
        activity.logger.info(f"Downloaded from MinIO to {local_path}")
        return local_path, True
    return filepath, False


def _convert_image_to_pdf(input_path: str, output_path: str) -> None:
    from PIL import Image, ImageOps

    with Image.open(input_path) as img:
        rgb_img = ImageOps.exif_transpose(img).convert("RGB")
        rgb_img.save(output_path, "PDF", resolution=300.0)


def _convert_office_to_pdf(input_path: str, output_dir: str) -> str:
    soffice_bin = shutil.which("soffice")
    if not soffice_bin:
        raise RuntimeError(
            "LibreOffice (soffice) is not installed. Install libreoffice to convert office files to PDF."
        )

    cmd = [
        soffice_bin,
        "--headless",
        "--nologo",
        "--nodefault",
        "--nolockcheck",
        "--nofirststartwizard",
        "--convert-to",
        "pdf",
        "--outdir",
        output_dir,
        input_path,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(
            f"Failed to convert office document to PDF: {result.stderr.strip() or result.stdout.strip()}"
        )

    output_pdf = Path(output_dir) / f"{Path(input_path).stem}.pdf"
    if not output_pdf.exists():
        raise RuntimeError("Office-to-PDF conversion finished but output PDF was not found")
    return str(output_pdf)


def _ensure_pdf_input(local_path: str) -> tuple[str, bool]:
    """
    Ensure the given file path points to a PDF.
    Returns (pdf_path, should_cleanup).
    """
    ext = Path(local_path).suffix.lower()
    if ext not in SUPPORTED_INPUT_EXTENSIONS:
        raise ValueError(
            f"Unsupported file extension '{ext}'. Supported extensions: {sorted(SUPPORTED_INPUT_EXTENSIONS)}"
        )

    if ext == ".pdf":
        return local_path, False

    work_dir = tempfile.mkdtemp(prefix="doc_convert_")
    output_pdf = Path(work_dir) / f"{Path(local_path).stem}.pdf"

    if ext in IMAGE_INPUT_EXTENSIONS:
        _convert_image_to_pdf(local_path, str(output_pdf))
    elif ext in OFFICE_INPUT_EXTENSIONS:
        converted = _convert_office_to_pdf(local_path, work_dir)
        if converted != str(output_pdf):
            shutil.move(converted, output_pdf)
    else:
        raise ValueError(f"Unsupported conversion path for extension: {ext}")

    activity.logger.info(f"Converted {local_path} -> {output_pdf}")
    return str(output_pdf), True


def _csv_to_pages(input_path: str, rows_per_page: int = 80) -> list[dict]:
    import csv

    def _open_reader():
        last_err = None
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                f = open(input_path, "r", encoding=enc, newline="")
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except Exception:
                    dialect = csv.excel
                return f, csv.reader(f, dialect=dialect)
            except Exception as e:
                last_err = e
        raise RuntimeError(f"Could not read CSV file: {last_err}")

    f, reader = _open_reader()
    try:
        rows = [[(c or "").strip() for c in r] for r in reader if any((c or "").strip() for c in r)]
    finally:
        f.close()

    if not rows:
        return [{
            "page_number": 1,
            "original_markdown": f"# {Path(input_path).name}\n\n(Empty CSV file)",
            "edited_markdown": None,
            "is_reviewed": False,
            "reviewer_notes": None,
        }]

    header = rows[0]
    body = rows[1:] if len(rows) > 1 else []

    pages: list[dict] = []
    page_num = 1
    for i in range(0, len(body), rows_per_page):
        batch = body[i:i + rows_per_page]
        lines = [f"# Spreadsheet Data: {Path(input_path).name}", "", f"Columns: {', '.join(header)}", ""]
        for row_idx, row in enumerate(batch, start=i + 1):
            pairs = []
            for col_i, val in enumerate(row):
                col = header[col_i] if col_i < len(header) and header[col_i] else f"col_{col_i+1}"
                pairs.append(f"{col}: {val}")
            if pairs:
                lines.append(f"- Row {row_idx}: " + " | ".join(pairs))
        pages.append({
            "page_number": page_num,
            "original_markdown": "\n".join(lines),
            "edited_markdown": None,
            "is_reviewed": False,
            "reviewer_notes": None,
        })
        page_num += 1
    return pages


def _xlsx_to_pages(input_path: str, rows_per_page: int = 80) -> list[dict]:
    from datetime import date, datetime
    from openpyxl import load_workbook

    def _cell_to_str(value) -> str:
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        return str(value).strip()

    pages: list[dict] = []
    page_num = 1
    wb = load_workbook(filename=input_path, data_only=True, read_only=True)

    for sheet in wb.worksheets:
        header: list[str] | None = None
        batch: list[tuple[int, list[str]]] = []
        seen_rows = 0

        def emit_page(rows_batch: list[tuple[int, list[str]]]) -> None:
            nonlocal page_num
            if not rows_batch:
                return
            assert header is not None
            lines = [
                f"# Spreadsheet Data: {Path(input_path).name}",
                "",
                f"Sheet: {sheet.title}",
                f"Columns: {', '.join(header)}",
                "",
            ]
            for excel_row_num, row in rows_batch:
                pairs = []
                for col_i, val in enumerate(row):
                    col = header[col_i] if col_i < len(header) and header[col_i] else f"col_{col_i+1}"
                    pairs.append(f"{col}: {val}")
                if pairs:
                    lines.append(f"- Row {excel_row_num}: " + " | ".join(pairs))
            pages.append({
                "page_number": page_num,
                "original_markdown": "\n".join(lines),
                "edited_markdown": None,
                "is_reviewed": False,
                "reviewer_notes": None,
            })
            page_num += 1

        for excel_row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [_cell_to_str(v) for v in row]
            if not any(values):
                continue
            seen_rows += 1

            if header is None:
                header = values
                continue

            batch.append((excel_row_num, values))
            if len(batch) >= rows_per_page:
                emit_page(batch)
                batch = []

        if header is not None:
            emit_page(batch)
        elif seen_rows == 0:
            pages.append({
                "page_number": page_num,
                "original_markdown": f"# Spreadsheet Data: {Path(input_path).name}\n\nSheet: {sheet.title}\n\n(Empty sheet)",
                "edited_markdown": None,
                "is_reviewed": False,
                "reviewer_notes": None,
            })
            page_num += 1

    wb.close()

    if not pages:
        return [{
            "page_number": 1,
            "original_markdown": f"# Spreadsheet Data: {Path(input_path).name}\n\n(Empty workbook)",
            "edited_markdown": None,
            "is_reviewed": False,
            "reviewer_notes": None,
        }]
    return pages


def _ocr_pdf(local_pdf_path: str) -> list[dict]:
    return run_ocr_pdf(local_pdf_path, clean_text, log=_temporal_log)


def _ocr_pdf_in_segments(
    local_pdf_path: str,
    segment_pages: int,
    on_segment_complete=None,
    completed_page_numbers: set[int] | None = None,
) -> list[dict]:
    return run_ocr_pdf_in_segments(
        local_pdf_path,
        segment_pages,
        clean_text,
        on_segment_complete=on_segment_complete,
        completed_page_numbers=completed_page_numbers,
        log=_temporal_log,
    )


def _temporal_log(level: str, message: str, *args) -> None:
    log_fn = getattr(activity.logger, level, activity.logger.info)
    log_fn(message, *args)


def _pdf_page_count(pdf_path: str) -> int:
    with fitz.open(pdf_path) as doc:
        return doc.page_count


def _validate_ocr_pages_for_pdf(pdf_path: str, pages: list[dict], *, filename: str) -> None:
    expected = _pdf_page_count(pdf_path)
    got = len(pages)
    if expected > 0 and got == 0:
        raise RuntimeError(
            f"OCR produced 0 pages for non-empty PDF ({expected} source page(s)): {filename}"
        )
    if expected > 0 and got != expected:
        raise RuntimeError(
            f"OCR page count mismatch for {filename}: source PDF has {expected} page(s), "
            f"OCR persisted {got}"
        )
    page_numbers = sorted(int(p.get("page_number") or 0) for p in pages if p.get("page_number"))
    expected_numbers = list(range(1, expected + 1))
    if expected > 0 and page_numbers != expected_numbers:
        raise RuntimeError(
            f"OCR page numbering mismatch for {filename}: expected {expected_numbers}, "
            f"got {page_numbers}"
        )


def _drop_degenerate_ocr_pages(pages: list[dict], *, filename: str) -> tuple[list[dict], list[int]]:
    kept: list[dict] = []
    dropped: list[int] = []
    for page in pages:
        text = page.get("original_markdown") or page.get("edited_markdown") or ""
        if is_degenerate_repetition(text):
            dropped.append(int(page.get("page_number") or 0))
            note = degenerate_ocr_note(text)
            activity.logger.warning(
                "Dropping degenerate OCR page %s from %s: %s",
                page.get("page_number"),
                filename,
                note,
            )
            continue
        kept.append(page)
    if not kept and dropped:
        raise RuntimeError(
            f"OCR produced only degenerate repetition output for {filename} "
            f"(pages {dropped})"
        )
    if dropped:
        activity.logger.warning(
            "Dropped %s degenerate OCR page(s) from %s: %s",
            len(dropped),
            filename,
            dropped,
        )
    return kept, dropped


def _finalize_ocr_pages(workflow_id: str, pages: list[dict], *, filename: str) -> list[dict]:
    """Drop degenerate OCR pages and persist the finalized SQLite page set."""
    from .. import db

    kept, dropped = _drop_degenerate_ocr_pages(pages, filename=filename)
    if dropped:
        removed = db.delete_pages(workflow_id, dropped)
        activity.logger.info(
            "Removed %s degenerate OCR page row(s) from SQLite for %s (pages %s)",
            removed,
            workflow_id,
            dropped,
        )
    # PDF segments are already persisted, so this is an idempotent upsert for
    # them. CSV/XLSX pages only exist in memory at this point and must be saved
    # here before downstream activities reload the page set from SQLite.
    db.save_pages(workflow_id, kept)
    return kept


def _live_progress_workflow_id(*candidates: object) -> str:
    """Best-effort document workflow_id for the Redis ticker."""
    for value in candidates:
        text = str(value or "").strip()
        if text:
            return text
    try:
        return str(activity.info().workflow_id or "").strip()
    except Exception:
        return ""


def _coalesce_translation_progress(
    *,
    event_phase: str,
    pages_completed: int,
    pages_total: int,
    prev: dict | None,
    sqlite_translated: int,
    document_page_count: int,
) -> tuple[int, int]:
    """Keep language-detection ticks from freezing the translation remaining-work bar."""
    phase = str(event_phase or "translation").strip().lower()
    completed = max(0, int(pages_completed or 0))
    total = max(0, int(pages_total or 0))
    if phase == "translation":
        return max(completed, int(sqlite_translated or 0)), total
    prev = prev or {}
    return (
        max(completed, int(prev.get("done") or 0)),
        max(total, int(prev.get("total") or 0), int(document_page_count or 0)),
    )


def _activity_heartbeat(payload: dict) -> None:
    """Send a Temporal heartbeat and publish the live-progress cache ticker."""
    try:
        from ..services.progress import publish_live_progress

        if isinstance(payload, dict) and not str(payload.get("workflow_id") or "").strip():
            wf_id = _live_progress_workflow_id()
            if wf_id:
                payload = {**payload, "workflow_id": wf_id}
        publish_live_progress(payload)
    except Exception:
        logging.debug("Could not publish live progress cache", exc_info=True)
    try:
        activity.heartbeat(payload)
    except RuntimeError as exc:
        # Unit tests call activities outside Temporal context.
        if "Not in activity context" not in str(exc):
            raise


def _build_chunks_from_pages(
    pages: list[dict],
    chunk_size: int = 450,
    chunk_overlap: int = 128,
    min_tokens: int = 100,
) -> list[dict]:
    raise RuntimeError("_build_chunks_from_pages is deprecated; use chunk_pages() via create_chunks_from_db")


class IngestWriteError(RuntimeError):
    """Raised when Marqo ingest fails after writing some records."""

    def __init__(
        self,
        message: str,
        *,
        records_ingested: int = 0,
        phase: str = "add_documents",
    ):
        super().__init__(message)
        self.records_ingested = max(0, int(records_ingested or 0))
        self.phase = phase or "add_documents"


def _verify_ingest_target_index(
    store,
    index_name: str,
    passage_fields: set[str],
    *,
    require_workflow_filter: bool = False,
) -> tuple[IndexSchemaReport, set[str]]:
    """Run all non-mutating index checks required before ingest/purge."""
    try:
        report = store.describe_index(index_name)
    except Exception as e:
        activity.logger.error(
            "Could not verify schema for index %s; aborting this attempt "
            "(transient error, letting Temporal retry): %s",
            index_name,
            e,
        )
        raise

    if not report.exists:
        return report, set(passage_fields)

    index_field_names = report.field_names or set()
    if require_workflow_filter and "workflow_id" not in index_field_names:
        raise RuntimeError(
            "Cannot safely replace Marqo records for this document: "
            f"index {index_name} does not expose workflow_id as a filterable field. "
            "Recreate/migrate the index with passage schema and reingest."
        )

    if report.missing_core or not report.has_passage_tensor:
        activity.logger.warning(
            "Index %s does not match the canonical passage schema "
            "(missing_core=%s, has_passage_tensor=%s). NOT recreating it — an "
            "existing index is never destroyed implicitly. Ingesting with the "
            "fields this index accepts; unsupported fields will be dropped. To "
            "migrate, recreate the index explicitly (admin endpoint with "
            "recreate_if_exists=true, or scripts/reset_marqo_index.py) and reingest.",
            index_name,
            report.missing_core,
            report.has_passage_tensor,
        )

    if index_field_names and not report.tensor_fields:
        raise RuntimeError(
            f"Index {index_name} declares no tensor fields; ingesting would store "
            "documents without embeddings (invisible to retrieval). Refusing to "
            "ingest. Recreate the index explicitly with the passage schema and reingest."
        )
    return report, index_field_names


async def _detect_and_translate_impl(
    pages: list[dict],
    target_language: str = "en",
    source_language: str | None = None,
    *,
    force_retranslate: bool = False,
    progress_callback: Callable[[dict], None] | None = None,
) -> list[dict]:
    del source_language
    config = load_translation_config(target_language=target_language)
    return await translate_pages(
        pages,
        target_language=target_language,
        config=config,
        log=activity.logger.info,
        force_retranslate=force_retranslate,
        progress_callback=progress_callback,
    )


# =============================================================================
# Activities
# =============================================================================

@activity.defn
async def run_ocr(filepath: str) -> list[dict]:
    """Run OCR on a supported file and return page dicts."""
    activity.logger.info(f"Running OCR on {filepath}")

    local_path, cleanup_local = _resolve_local_path(filepath)
    ext = Path(local_path).suffix.lower()
    pdf_path = local_path
    cleanup_pdf_dir = False

    try:
        if ext in DELIMITED_INPUT_EXTENSIONS:
            pages = _csv_to_pages(local_path)
            activity.logger.info(f"CSV parsed into {len(pages)} pages")
            return pages
        if ext in NATIVE_SPREADSHEET_EXTENSIONS:
            pages = _xlsx_to_pages(local_path)
            activity.logger.info(f"XLSX parsed into {len(pages)} pages")
            return pages
        pdf_path, cleanup_pdf_dir = _ensure_pdf_input(local_path)
        return _ocr_pdf(pdf_path)
    finally:
        if cleanup_pdf_dir:
            try:
                shutil.rmtree(Path(pdf_path).parent, ignore_errors=True)
            except Exception:
                pass
        if cleanup_local and os.path.exists(local_path):
            os.remove(local_path)


@activity.defn
async def run_ocr_and_store(
    workflow_id: str,
    filepath: str,
    force_redo: bool = False,
    discard_edits: bool = False,
    retry_job_id: int | None = None,
) -> dict:
    """Run OCR and persist pages to SQLite to avoid large Temporal payloads.

    ``force_redo`` clears existing pages for this workflow before OCR so segment
    resume cannot skip stale text (#123). Default ``False`` preserves crash resume
    for Temporal retries and ``POST /retry-ocr`` without force.

    When forcing, operator OCR edits are snapshotted and restored unless
    ``discard_edits`` is true. Prior MinIO ``ocr_pages_json`` artifacts are left
    in place; a new export is written after this run.
    """
    from .. import db

    local_path, cleanup_local = _resolve_local_path(filepath)
    ext = Path(local_path).suffix.lower()
    source_type, canonical_input_type = _source_type_from_path(local_path)
    original_filename = Path(local_path).name
    normalized_path = local_path
    cleanup_normalized = False
    segment_pages = max(
        1,
        int(os.environ.get("OCR_SEGMENT_PAGES", "5")),
    )
    latest_job = None
    if retry_job_id:
        latest_job = db.get_document_job(int(retry_job_id))
        if not latest_job:
            activity.logger.warning(
                "Retry OCR job id %s was not found for %s; falling back to latest job lookup.",
                retry_job_id,
                workflow_id,
            )
    if not latest_job:
        latest_job = db.get_latest_document_job(workflow_id)
    job_id = latest_job["id"] if latest_job else None
    job_config: dict = {}
    if latest_job and latest_job.get("config_json"):
        try:
            job_config = json.loads(latest_job["config_json"]) or {}
        except Exception:
            job_config = {}
    edit_snapshot: dict[int, dict] = {}

    try:
        if force_redo:
            force_state = job_config.get("force_ocr_state") if isinstance(job_config, dict) else None
            if force_state and force_state.get("initialized"):
                if not discard_edits:
                    edit_snapshot = force_state.get("edit_snapshot") or {}
                activity.logger.info(
                    "Force re-OCR retry for %s: reusing durable init "
                    "(discard_edits=%s, edit_snapshot_pages=%s)",
                    workflow_id,
                    discard_edits,
                    len(edit_snapshot),
                )
            else:
                if not discard_edits:
                    edit_snapshot = db.snapshot_page_ocr_edits(workflow_id)
                removed = db.delete_pages(workflow_id)
                activity.logger.info(
                    "Force re-OCR for %s: cleared %s page row(s) "
                    "(discard_edits=%s, edit_snapshot_pages=%s)",
                    workflow_id,
                    removed,
                    discard_edits,
                    len(edit_snapshot),
                )
                db.update_document_fields(workflow_id, page_count=0)
                if latest_job:
                    next_config = dict(job_config)
                    next_config["force_ocr_state"] = {
                        "initialized": True,
                        "discard_edits": bool(discard_edits),
                        "edit_snapshot": edit_snapshot if not discard_edits else {},
                    }
                    db.update_document_job(job_id, config_json=next_config)
                    job_config = next_config

        if ext in DELIMITED_INPUT_EXTENSIONS:
            pages = _csv_to_pages(local_path)
        elif ext in NATIVE_SPREADSHEET_EXTENSIONS:
            pages = _xlsx_to_pages(local_path)
        else:
            normalized_path, cleanup_normalized = _ensure_pdf_input(local_path)
            # After force_redo the set is empty; otherwise resume skips saved pages.
            saved_page_numbers = set(db.get_saved_page_numbers(workflow_id))
            loop = asyncio.get_running_loop()

            def persist_segment(segment_pages_result: list[dict], total_pages: int) -> None:
                db.save_pages(workflow_id, segment_pages_result)
                current_saved = len(saved_page_numbers.union({p["page_number"] for p in segment_pages_result}))
                saved_page_numbers.update(p["page_number"] for p in segment_pages_result)
                db.update_document_fields(workflow_id, page_count=current_saved)
                payload = {
                    "workflow_id": workflow_id,
                    "pages_saved": current_saved,
                    "total_pages": total_pages,
                    "stage": "ocr",
                    "phase": "ocr",
                    "done": current_saved,
                    "total": total_pages,
                    "unit": "pages",
                    "updated_at": datetime.utcnow().isoformat(),
                }
                loop.call_soon_threadsafe(_activity_heartbeat, payload)
                activity.logger.info(
                    "Persisted OCR segment for %s: %s/%s pages saved",
                    workflow_id,
                    current_saved,
                    total_pages,
                )

            total_pages_hint = _pdf_page_count(normalized_path)
            _activity_heartbeat(
                {
                    "workflow_id": workflow_id,
                    "pages_saved": len(saved_page_numbers),
                    "total_pages": total_pages_hint,
                    "stage": "ocr",
                    "phase": "ocr",
                    "done": len(saved_page_numbers),
                    "total": total_pages_hint,
                    "unit": "pages",
                    "updated_at": datetime.utcnow().isoformat(),
                }
            )

            pages = await asyncio.to_thread(
                _ocr_pdf_in_segments,
                normalized_path,
                segment_pages=segment_pages,
                on_segment_complete=persist_segment,
                completed_page_numbers=saved_page_numbers,
            )

            pages = db.get_pages(workflow_id)
            _validate_ocr_pages_for_pdf(normalized_path, pages, filename=original_filename)

        pages = _finalize_ocr_pages(workflow_id, pages, filename=original_filename)
        if edit_snapshot:
            restored = db.restore_page_ocr_edits(workflow_id, edit_snapshot)
            activity.logger.info(
                "Restored OCR edits for %s on %s/%s snapshotted page(s)",
                workflow_id,
                restored,
                len(edit_snapshot),
            )
            pages = db.get_pages(workflow_id)

        # Tenant prefix for new artifact writes (from the durable SQLite row).
        doc_instance = (db.get_document(workflow_id) or {}).get("instance")

        normalized_filename = _normalized_filename(original_filename, canonical_input_type)
        norm_artifact_type = _normalized_artifact_type(canonical_input_type)
        normalized_uri, normalized_size, normalized_mime = _upload_file_to_minio(
            normalized_path,
            workflow_id,
            norm_artifact_type,
            normalized_filename,
            instance=doc_instance,
        )
        normalized_artifact_id = db.add_document_artifact(
            workflow_id=workflow_id,
            job_id=job_id,
            artifact_type=norm_artifact_type,
            stage="ocr_processing",
            storage_uri=normalized_uri,
            mime_type=normalized_mime,
            filename=normalized_filename,
            size_bytes=normalized_size,
            metadata={"source_filepath": filepath, "canonical_input_type": canonical_input_type},
        )

        if filepath.startswith("minio://"):
            original_uri = filepath
            original_size = None
            original_mime = mimetypes.guess_type(original_filename)[0] or "application/octet-stream"
        else:
            original_uri, original_size, original_mime = _upload_file_to_minio(
                local_path,
                workflow_id,
                "original_upload",
                original_filename,
                instance=doc_instance,
            )
        original_artifact_id = db.add_document_artifact(
            workflow_id=workflow_id,
            job_id=job_id,
            artifact_type="original_upload",
            stage="registered",
            storage_uri=original_uri,
            mime_type=original_mime,
            filename=original_filename,
            size_bytes=original_size,
            metadata={"source_filepath": filepath},
        )

        pages_json_path = _write_json_temp(pages)
        try:
            pages_uri, pages_size, pages_mime = _upload_file_to_minio(
                pages_json_path, workflow_id, "ocr_pages_json", "pages.json", instance=doc_instance
            )
        finally:
            if os.path.exists(pages_json_path):
                os.remove(pages_json_path)

        db.add_document_artifact(
            workflow_id=workflow_id,
            job_id=job_id,
            artifact_type="ocr_pages_json",
            stage="ocr_review",
            storage_uri=pages_uri,
            mime_type=pages_mime,
            filename="pages.json",
            size_bytes=pages_size,
            metadata={"page_count": len(pages)},
        )

        db.update_document_fields(
            workflow_id,
            page_count=len(pages),
            source_type=source_type,
            canonical_input_type=canonical_input_type,
            original_artifact_id=original_artifact_id,
            normalized_artifact_id=normalized_artifact_id,
        )
        if force_redo and latest_job and "force_ocr_state" in job_config:
            next_config = dict(job_config)
            next_config.pop("force_ocr_state", None)
            db.update_document_job(job_id, config_json=next_config)
        return {"page_count": len(pages), "normalized_artifact_id": normalized_artifact_id}
    finally:
        if cleanup_normalized and os.path.exists(normalized_path):
            shutil.rmtree(Path(normalized_path).parent, ignore_errors=True)
        if cleanup_local and os.path.exists(local_path):
            os.remove(local_path)


@activity.defn
async def create_chunks(
    pages: list[dict],
    chunk_size: int = 450,
    chunk_overlap: int = 128,
    min_tokens: int = 100,
) -> list[dict]:
    """Create chunks from pages with page range tracking."""
    activity.logger.info(f"Creating chunks from {len(pages)} pages")
    config = load_chunking_config(chunk_size=chunk_size, chunk_overlap=chunk_overlap, min_tokens=min_tokens)
    result = await chunk_pages(pages, config)
    chunks = []
    for idx, chunk in enumerate(result.chunks, 1):
        chunks.append(
            {
                "chunk_number": idx,
                "original_text": chunk.text,
                "edited_text": None,
                "token_count": chunk.token_count,
                "page_start": chunk.page_start,
                "page_end": chunk.page_end,
                "source_page_numbers_json": json.dumps(chunk.source_page_numbers),
                "source_spans_json": json.dumps(chunk.source_spans),
                "section_title": chunk.section_title,
                "content_type": chunk.content_type,
                "is_reference": chunk.is_reference,
                "chunking_provider": result.provider,
                "chunking_model": result.model,
                "chunking_config_json": result.config.to_json(),
                "chunking_run_id": "",
                "chunk_version": 1,
                "is_reviewed": False,
                "is_excluded": False,
                "reviewer_notes": None,
            }
        )
    activity.logger.info(f"Created {len(chunks)} chunks with page tracking")
    return chunks


@activity.defn
async def create_chunks_from_db(
    workflow_id: str,
    chunk_size: int = 450,
    chunk_overlap: int = 128,
    min_tokens: int = 100,
    retry_job_id: int | None = None,
) -> dict:
    """Create chunks from persisted pages and persist chunks in SQLite.

    ``retry_job_id`` is the bound ``document_jobs`` row for this run (initial
    pipeline or chunking retry). When unset, fall back to the latest job.
    """
    from .. import db

    pages = db.get_pages(workflow_id)
    activity.logger.info(f"Creating chunks from DB pages for {workflow_id}: {len(pages)} pages")
    config = load_chunking_config(chunk_size=chunk_size, chunk_overlap=chunk_overlap, min_tokens=min_tokens)
    checkpoint_windows = max(1, int(os.environ.get("CHUNKING_CHECKPOINT_WINDOWS", "25")))
    checkpoint_min_pages = max(1, int(os.environ.get("CHUNKING_CHECKPOINT_MIN_PAGES", "500")))
    latest_job = None
    if retry_job_id:
        latest_job = db.get_document_job(int(retry_job_id))
        if not latest_job:
            activity.logger.warning(
                "Bound chunking job id %s was not found for %s; falling back to latest job lookup.",
                retry_job_id,
                workflow_id,
            )
    if not latest_job:
        latest_job = db.get_latest_document_job(workflow_id)
    base_job_config: dict = {}
    if latest_job and latest_job.get("config_json"):
        try:
            base_job_config = json.loads(latest_job["config_json"]) or {}
        except Exception:
            base_job_config = {}
    job_config = dict(base_job_config)

    checkpoint_mode = (
        is_llm_grouping_provider(config.provider)
        and len(pages) >= checkpoint_min_pages
        and checkpoint_windows > 0
    )
    if checkpoint_mode:
        # Avoid provider fallback while checkpointing: retries should resume from
        # persisted windows, not silently switch provider semantics mid-run.
        config = replace(config, fallback_provider=config.provider)

    window_size = max(1, int(getattr(config, "page_window_size", 1) or 1))
    total_windows = max(1, (len(pages) + window_size - 1) // window_size) if pages else 0
    chunking_run_id = f"chunk-{workflow_id}-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    chunk_version = db.next_chunk_version(workflow_id)
    windows_completed = 0
    chunks_persisted = 0
    next_chunk_number = 1
    resume_page_offset = 0
    next_page_offset = 0
    last_chunk_norm = ""
    last_chunk_end_page = 0
    checkpoint_warnings: list[str] = []
    pages_for_chunking = pages
    pending_checkpoint_rows: list[dict] = []
    pending_windows_since_flush = 0

    checkpoint_state = job_config.get("chunk_checkpoint") if isinstance(job_config, dict) else None
    can_resume = (
        checkpoint_mode
        and isinstance(checkpoint_state, dict)
        and checkpoint_state.get("status") in {"running", "interrupted", "completed"}
        and checkpoint_state.get("provider") == config.provider
        and int(checkpoint_state.get("pages_total") or 0) == len(pages)
        and int(checkpoint_state.get("total_windows") or 0) == total_windows
        and int(checkpoint_state.get("windows_completed") or 0) > 0
        and checkpoint_state.get("chunking_run_id")
    )
    if can_resume:
        chunking_run_id = str(checkpoint_state.get("chunking_run_id"))
        chunk_version = int(checkpoint_state.get("chunk_version") or chunk_version)
        windows_completed = int(checkpoint_state.get("windows_completed") or 0)
        chunks_persisted = int(checkpoint_state.get("chunks_persisted") or 0)
        next_chunk_number = chunks_persisted + 1
        resume_page_offset = int(checkpoint_state.get("next_page_offset") or (windows_completed * window_size))
        next_page_offset = min(len(pages), max(0, resume_page_offset))
        pages_for_chunking = pages[resume_page_offset:]
        prior_chunks = db.get_chunks(workflow_id, include_excluded=True)
        if prior_chunks:
            last_chunk = prior_chunks[-1]
            last_chunk_norm = normalize_text(last_chunk.get("edited_text") or last_chunk.get("original_text") or "")
            last_chunk_end_page = int(last_chunk.get("page_end") or 0)
    elif checkpoint_mode:
        db.reset_chunks_for_checkpoint(workflow_id)
        next_page_offset = 0

    def _checkpoint_snapshot(*, status: str = "running") -> dict:
        return {
            "status": status,
            "provider": config.provider,
            "chunking_run_id": chunking_run_id,
            "chunk_version": chunk_version,
            "pages_total": len(pages),
            "total_windows": total_windows,
            "windows_completed": windows_completed,
            "chunks_persisted": chunks_persisted,
            "next_page_offset": next_page_offset,
            "checkpoint_windows": checkpoint_windows,
            "updated_at": datetime.utcnow().isoformat(),
        }

    def _persist_job_config(progress: dict | None = None, checkpoint_status: str | None = None) -> None:
        if not latest_job:
            return
        if progress is not None:
            job_config["chunking_progress"] = progress
        if checkpoint_mode:
            if checkpoint_status is not None:
                job_config["chunk_checkpoint"] = _checkpoint_snapshot(status=checkpoint_status)
            elif "chunk_checkpoint" not in job_config:
                job_config["chunk_checkpoint"] = _checkpoint_snapshot(status="running")
        db.update_document_job(latest_job["id"], config_json=job_config)

    def _flush_checkpoint_rows() -> None:
        nonlocal chunks_persisted, pending_windows_since_flush, pending_checkpoint_rows
        if not pending_checkpoint_rows:
            return
        db.append_chunk_checkpoint(workflow_id, pending_checkpoint_rows)
        chunks_persisted += len(pending_checkpoint_rows)
        pending_checkpoint_rows = []
        pending_windows_since_flush = 0

    async def _persist_chunking_progress(event: dict) -> None:
        nonlocal windows_completed, chunks_persisted, next_chunk_number, last_chunk_norm, last_chunk_end_page
        nonlocal next_page_offset, pending_windows_since_flush, pending_checkpoint_rows
        pages_total_local = int(event.get("pages_total") or len(pages_for_chunking) or 0)
        pages_processed_local = int(event.get("pages_processed") or 0)
        checkpoint_rows: list[dict] = []
        if checkpoint_mode and bool(event.get("window_succeeded")):
            pages_processed_absolute = min(len(pages), resume_page_offset + pages_processed_local)
            next_page_offset = max(next_page_offset, pages_processed_absolute)
            for candidate in event.get("checkpoint_window_chunks") or []:
                candidate_text = str(candidate.get("text") or "")
                candidate_norm = normalize_text(candidate_text)
                page_start = int(candidate.get("page_start") or 1)
                page_end = int(candidate.get("page_end") or page_start)
                if (
                    last_chunk_norm
                    and candidate_norm
                    and candidate_norm == last_chunk_norm
                    and page_start <= (last_chunk_end_page + 1)
                ):
                    checkpoint_warnings.append(
                        "Dropped adjacent checkpoint chunk with identical text on pages "
                        f"{page_start}-{page_end}"
                    )
                    continue
                checkpoint_rows.append(
                    {
                        "chunk_number": next_chunk_number,
                        "original_text": candidate_text,
                        "edited_text": None,
                        "token_count": int(candidate.get("token_count") or 0),
                        "page_start": page_start,
                        "page_end": page_end,
                        "source_page_numbers_json": json.dumps(candidate.get("source_page_numbers") or []),
                        "source_spans_json": json.dumps(candidate.get("source_spans") or []),
                        "section_title": candidate.get("section_title"),
                        "content_type": candidate.get("content_type"),
                        "is_reference": bool(candidate.get("is_reference")),
                        "chunking_provider": config.provider,
                        "chunking_model": config.model,
                        "chunking_config_json": config.to_json(),
                        "chunking_run_id": chunking_run_id,
                        "chunk_version": chunk_version,
                        "is_reviewed": False,
                        "is_excluded": False,
                        "reviewer_notes": None,
                    }
                )
                next_chunk_number += 1
                if candidate_norm:
                    last_chunk_norm = candidate_norm
                last_chunk_end_page = max(last_chunk_end_page, page_end)
            if checkpoint_rows:
                pending_checkpoint_rows.extend(checkpoint_rows)
            pending_windows_since_flush += 1
            if pending_windows_since_flush >= checkpoint_windows:
                _flush_checkpoint_rows()
            windows_completed = min(total_windows, windows_completed + 1)

        if checkpoint_mode:
            pages_processed = next_page_offset
            pages_total = len(pages)
            chunks_emitted = chunks_persisted + len(pending_checkpoint_rows)
            raw_percent = (windows_completed / total_windows * 100.0) if total_windows else 100.0
        else:
            pages_total = int(event.get("pages_total") or len(pages) or 0)
            pages_processed = int(event.get("pages_processed") or 0)
            chunks_emitted = int(event.get("chunks_emitted") or 0)
            raw_percent = float(event.get("percent") or 0.0)
        percent = max(0.0, min(100.0, raw_percent))
        updated_at = datetime.utcnow().isoformat()
        progress = {
            "status": "running" if percent < 100.0 else "completed",
            "provider": event.get("provider") or config.provider,
            "pages_processed": pages_processed,
            "pages_total": pages_total,
            "chunks_emitted": chunks_emitted,
            "percent": round(percent, 2),
            "updated_at": updated_at,
        }
        checkpoint_status = "running"
        if checkpoint_mode and windows_completed >= total_windows:
            checkpoint_status = "completed"
        _activity_heartbeat(
            {
                "workflow_id": workflow_id,
                "stage": "chunking",
                "phase": "chunking",
                "done": pages_processed,
                "total": pages_total,
                "unit": "pages",
                "updated_at": updated_at,
                "pages_processed": pages_processed,
                "pages_total": pages_total,
                "chunks_emitted": chunks_emitted,
                "percent": progress["percent"],
            }
        )
        _persist_job_config(progress=progress, checkpoint_status=checkpoint_status)

    await _persist_chunking_progress(
        {
            "provider": config.provider,
            "pages_processed": 0,
            "pages_total": len(pages_for_chunking),
            "chunks_emitted": 0,
            "percent": 0.0,
        }
    )

    try:
        if checkpoint_mode and not pages_for_chunking:
            # LLM windows already finished on a prior attempt; continue with
            # export/finalization without re-running expensive chunk generation.
            result = await chunk_pages([], config, progress_callback=None)
        else:
            result = await chunk_pages(pages_for_chunking, config, progress_callback=_persist_chunking_progress)
            if checkpoint_mode:
                _flush_checkpoint_rows()
    except Exception:
        if checkpoint_mode:
            _flush_checkpoint_rows()
            _persist_job_config(checkpoint_status="interrupted")
        raise

    checkpoint_finalize_warnings: list[str] = []
    if checkpoint_mode:
        persisted_rows = db.get_chunks(workflow_id, include_excluded=True)
        finalized_candidates: list[ChunkCandidate] = []
        final_last_norm = ""
        final_last_end_page = 0
        for row in persisted_rows:
            text = str(row.get("edited_text") or row.get("original_text") or "")
            norm = normalize_text(text)
            page_start = int(row.get("page_start") or 1)
            page_end = int(row.get("page_end") or page_start)
            if (
                final_last_norm
                and norm
                and norm == final_last_norm
                and page_start <= (final_last_end_page + 1)
            ):
                checkpoint_finalize_warnings.append(
                    "Dropped adjacent checkpoint chunk with identical text on pages "
                    f"{page_start}-{page_end}"
                )
                continue
            finalized_candidates.append(
                ChunkCandidate(
                    text=text,
                    page_start=page_start,
                    page_end=page_end,
                    source_page_numbers=_json_list(row.get("source_page_numbers_json")),
                    source_spans=_json_list(row.get("source_spans_json")),
                    token_count=int(row.get("token_count") or 0),
                    section_title=row.get("section_title") or "",
                    content_type=row.get("content_type") or "body",
                    is_reference=bool(row.get("is_reference")),
                )
            )
            if norm:
                final_last_norm = norm
            final_last_end_page = max(final_last_end_page, page_end)

        # Checkpoint rows are raw provider candidates persisted per window, before
        # chunk_pages could apply the post-chunk guards. Apply them once here, on
        # the whole reconstructed set, so a resumed run stores the same limits a
        # single-pass run would (#115).
        finalized_result = enforce_chunk_limits(
            ChunkingResult(
                chunks=finalized_candidates,
                provider=config.provider,
                model=config.model,
                config=config,
            )
        )
        checkpoint_finalize_warnings.extend(finalized_result.warnings)

        chunks = [
            {
                "chunk_number": idx,
                "original_text": candidate.text,
                "edited_text": None,
                "token_count": candidate.token_count,
                "page_start": candidate.page_start,
                "page_end": candidate.page_end,
                "source_page_numbers_json": json.dumps(candidate.source_page_numbers),
                "source_spans_json": json.dumps(candidate.source_spans),
                "section_title": candidate.section_title,
                "content_type": candidate.content_type,
                "is_reference": candidate.is_reference,
                "chunking_provider": config.provider,
                "chunking_model": config.model,
                "chunking_config_json": config.to_json(),
                "chunking_run_id": chunking_run_id,
                "chunk_version": chunk_version,
                "is_reviewed": False,
                "is_excluded": False,
                "reviewer_notes": None,
            }
            for idx, candidate in enumerate(finalized_result.chunks, 1)
        ]
        db.save_chunks(workflow_id, chunks)
    else:
        chunks = []
        for idx, chunk in enumerate(result.chunks, 1):
            chunks.append(
                {
                    "chunk_number": idx,
                    "original_text": chunk.text,
                    "edited_text": None,
                    "token_count": chunk.token_count,
                    "page_start": chunk.page_start,
                    "page_end": chunk.page_end,
                    "source_page_numbers_json": json.dumps(chunk.source_page_numbers),
                    "source_spans_json": json.dumps(chunk.source_spans),
                    "section_title": chunk.section_title,
                    "content_type": chunk.content_type,
                    "is_reference": chunk.is_reference,
                    "chunking_provider": result.provider,
                    "chunking_model": result.model,
                    "chunking_config_json": result.config.to_json(),
                    "chunking_run_id": chunking_run_id,
                    "chunk_version": chunk_version,
                    "is_reviewed": False,
                    "is_excluded": False,
                    "reviewer_notes": None,
                }
            )
        db.save_chunks(workflow_id, chunks)

    if checkpoint_mode:
        _persist_job_config(checkpoint_status="completed")

    artifact_stats = dict(result.stats or {})
    if checkpoint_mode:
        artifact_stats["chunk_count"] = len(chunks)
        artifact_stats["page_count"] = len(pages)

    chunks_instance = (db.get_document(workflow_id) or {}).get("instance")
    chunks_json_path = _write_json_temp(chunks)
    try:
        chunks_uri, chunks_size, chunks_mime = _upload_file_to_minio(
            chunks_json_path, workflow_id, "chunk_json_export", "chunks.json", instance=chunks_instance
        )
    finally:
        if os.path.exists(chunks_json_path):
            os.remove(chunks_json_path)
    db.add_document_artifact(
        workflow_id=workflow_id,
        job_id=latest_job["id"] if latest_job else None,
        artifact_type="chunk_json_export",
        stage="chunk_review",
        storage_uri=chunks_uri,
        mime_type=chunks_mime,
        filename="chunks.json",
        size_bytes=chunks_size,
        metadata={
            "chunk_count": len(chunks),
            "chunking_provider": result.provider,
            "chunking_model": result.model,
            "chunking_run_id": chunking_run_id,
            "chunking_config": json.loads(result.config.to_json()),
            "warnings": result.warnings + checkpoint_warnings + checkpoint_finalize_warnings,
            "stats": artifact_stats,
        },
    )
    # Reconcile the document row immediately after chunk persistence so SQLite
    # remains truthful even if the workflow fails before its final state update.
    db.reconcile_materialized_state(workflow_id)
    await _persist_chunking_progress(
        {
            "provider": result.provider,
            "pages_processed": len(pages),
            "pages_total": len(pages),
            "chunks_emitted": len(chunks),
            "percent": 100.0,
        }
    )
    return {"chunk_count": len(chunks)}


@activity.defn
async def prepare_for_ingestion(
    document_id: str,
    filename: str,
    chunks: list[dict],
    workflow_id: str | None = None,
    name_gu: str = None,
    name_en: str = None,
    description: str = None,
) -> list[dict]:
    """Prepare chunks for Marqo ingestion."""
    from .. import db

    activity.logger.info(f"Preparing {len(chunks)} chunks for ingestion")
    doc = db.get_document(workflow_id) if workflow_id else None
    records = _prepare_records(
        document_id,
        filename,
        chunks,
        workflow_id=workflow_id,
        name_gu=name_gu,
        name_en=name_en,
        description=description,
        instance=(doc or {}).get("instance"),
    )
    activity.logger.info(f"Prepared {len(records)} records")
    return records


@activity.defn
async def ingest_to_marqo(
    records: list[dict],
    marqo_url: str = None,
    index_name: str = "documents-index",
    batch_size: int = 10,
) -> dict:
    """Ingest records to Marqo.

    Every Marqo call goes through :mod:`pipeline.vector_store`. What stays here is
    ingest POLICY — the four decisions below about an index's lifecycle, each of
    which the adapter deliberately refuses to make on our behalf.

    ``marqo_url`` is dead (always ``""``) but kept: Temporal replays this
    activity's arguments positionally out of workflow history.
    """
    del marqo_url  # SSRF: the endpoint comes from the environment, never a caller.

    store = get_vector_store()
    activity.logger.info(f"Ingesting {len(records)} records to Marqo at {store.url}")

    passage_fields = passage_schema_field_names()
    report, index_field_names = _verify_ingest_target_index(store, index_name, passage_fields)

    if not report.exists:
        # Provisioning a NEW index is the only index-lifecycle action this activity
        # takes. It is safe: there is nothing to lose.
        store.create_index(index_name, passage_index_settings())
        activity.logger.info(f"Created index: {index_name} (passage schema)")
        index_field_names = set(passage_fields)

    ingest_workflow_id = _live_progress_workflow_id(
        *(str((row or {}).get("workflow_id") or "") for row in (records or [])[:1])
    )
    records = project_records(records, passage_fields, index_field_names)
    records_ingested_so_far = 0
    batch_count = 0
    rows_seen = 0

    def _report_batch(errors: list[dict], result: dict) -> None:
        nonlocal records_ingested_so_far, batch_count, rows_seen
        batch_count += 1
        items = list(result.get("items") or []) if isinstance(result, dict) else []
        rows_seen += len(items)
        batch_success = 0
        for item in items:
            if item.get("status") == 200:
                batch_success += 1
        records_ingested_so_far += batch_success
        updated_at = datetime.utcnow().isoformat()
        _activity_heartbeat(
            {
                "workflow_id": ingest_workflow_id,
                "stage": "ingest",
                "phase": "ingest",
                "done": rows_seen,
                "total": len(records),
                "unit": "chunks",
                "updated_at": updated_at,
                "batch": batch_count,
                "rows_seen": rows_seen,
                "rows_total": len(records),
            }
        )
        if not result.get("errors"):
            return
        activity.logger.error(
            "Marqo add_documents reported errors. First few: %s. Full result keys: %s",
            errors[:5],
            list(result.keys()),
        )
        if errors:
            raise IngestWriteError(
                f"Marqo add_documents failed for {len(errors)} doc(s). First error: {errors[0]}",
                records_ingested=records_ingested_so_far,
            )

    try:
        store.add_documents(index_name, records, batch_size=batch_size, on_batch=_report_batch)
    except IngestWriteError:
        raise
    except Exception as exc:
        # Transport/request failures can happen after earlier batches succeeded.
        # Preserve truthful partial progress for caller-side status accounting.
        raise IngestWriteError(
            f"Marqo add_documents request failed: {exc}",
            records_ingested=records_ingested_so_far,
        ) from exc

    try:
        stats = store.get_stats(index_name)
    except Exception as exc:
        # All add batches already succeeded at this point. Preserve truthful
        # progress accounting for caller-side index status updates.
        raise IngestWriteError(
            f"Marqo post-ingest stats lookup failed: {exc}",
            records_ingested=len(records),
            phase="post_ingest_stats",
        ) from exc
    activity.logger.info(f"Ingestion complete: {stats}")

    return {
        "records_ingested": len(records),
        "index_stats": stats,
        # Truthful per-index: an older index may not carry the prefixed tensor field.
        "supports_prefixed_tensor_field": ("text_for_embedding" in index_field_names)
        if index_field_names
        else True,
    }


@activity.defn
async def ingest_document_from_db(
    workflow_id: str,
    document_id: str,
    filename: str,
    marqo_url: str = None,
    index_name: str = "documents-index",
    batch_size: int = 10,
) -> dict:
    """Prepare and ingest chunks directly from SQLite by workflow_id."""
    from .. import db

    chunks = db.get_chunks(workflow_id, include_excluded=True)
    doc = db.get_document(workflow_id)
    # Prefer an explicit physical target when it is registered to this document's
    # tenant (Indexes-scoped bulk reindex). Otherwise registry-resolve / ensure
    # the tenant's own index — never another tenant's physical index. The Marqo
    # index itself is created by ingest_to_marqo below if it doesn't exist.
    index_name = db.resolve_ingest_index_name(
        (doc or {}).get("instance"),
        (doc or {}).get("index"),
        requested_index_name=index_name,
    )
    records = _prepare_records(
        document_id,
        filename,
        chunks,
        workflow_id=workflow_id,
        instance=(doc or {}).get("instance"),
    )
    payload_path = _write_json_temp(records)
    try:
        payload_uri, payload_size, payload_mime = _upload_file_to_minio(
            payload_path, workflow_id, "marqo_payload_export", "marqo_payload.json",
            instance=(doc or {}).get("instance"),
        )
    finally:
        if os.path.exists(payload_path):
            os.remove(payload_path)
    latest_job = db.get_latest_document_job(workflow_id)
    db.add_document_artifact(
        workflow_id=workflow_id,
        job_id=latest_job["id"] if latest_job else None,
        artifact_type="marqo_payload_export",
        stage="ingesting",
        storage_uri=payload_uri,
        mime_type=payload_mime,
        filename="marqo_payload.json",
        size_bytes=payload_size,
        metadata={"record_count": len(records), "index_name": index_name},
    )
    store = get_vector_store()
    # Run all non-mutating checks before purge so a target that cannot accept
    # writes never gets wiped first.
    _verify_ingest_target_index(
        store,
        index_name,
        passage_schema_field_names(),
        require_workflow_filter=True,
    )
    db.upsert_document_index_status(
        workflow_id=workflow_id,
        index_name=index_name,
        marqo_doc_id=document_id,
        chunk_count_indexed=0,
        last_verified_at=datetime.utcnow().isoformat(),
        schema_version="passage-v1",
        status="replacing",
        details={"record_count_expected": len(records)},
    )
    try:
        try:
            purge_result = store.delete_document(
                document_id,
                index_name,
                workflow_id=workflow_id,
            )
            if purge_result.get("error"):
                raise RuntimeError(purge_result["error"])
        except Exception as exc:
            # Tag the phase at the point it is known rather than inferring it
            # from the message later; nothing has been written yet here.
            raise IngestWriteError(
                f"Marqo purge before ingest failed for workflow {workflow_id}: {exc}",
                records_ingested=0,
                phase="purge",
            ) from exc
        activity.logger.info(
            "Purged %s Marqo record(s) for workflow %s before ingest (doc_id=%s)",
            purge_result.get("deleted", 0),
            workflow_id,
            document_id,
        )
        result = await ingest_to_marqo(
            records, marqo_url=marqo_url, index_name=index_name, batch_size=batch_size
        )
    except IngestWriteError as exc:
        db.upsert_document_index_status(
            workflow_id=workflow_id,
            index_name=index_name,
            marqo_doc_id=document_id,
            chunk_count_indexed=exc.records_ingested,
            last_verified_at=datetime.utcnow().isoformat(),
            schema_version="passage-v1",
            status="index_failed",
            details={"error": str(exc), "phase": exc.phase},
        )
        raise
    except Exception as exc:
        db.upsert_document_index_status(
            workflow_id=workflow_id,
            index_name=index_name,
            marqo_doc_id=document_id,
            chunk_count_indexed=0,
            last_verified_at=datetime.utcnow().isoformat(),
            schema_version="passage-v1",
            status="index_failed",
            details={"error": str(exc), "phase": "ingest"},
        )
        raise
    db.upsert_document_index_status(
        workflow_id=workflow_id,
        index_name=index_name,
        marqo_doc_id=document_id,
        chunk_count_indexed=result.get("records_ingested", 0),
        last_indexed_at=datetime.utcnow().isoformat(),
        last_verified_at=datetime.utcnow().isoformat(),
        schema_version="passage-v1",
        status="indexed",
        details=result.get("index_stats"),
    )
    return result


@activity.defn
async def update_document_state(
    workflow_id: str,
    stage: str,
    page_count: int = 0,
    chunk_count: int = 0,
    error_message: str = None,
) -> dict:
    """Update document state in SQLite."""
    from .. import db

    activity.logger.info(f"Updating state for {workflow_id}: stage={stage}")
    db.update_document_stage(
        workflow_id=workflow_id,
        stage=stage,
        page_count=page_count,
        chunk_count=chunk_count,
        error_message=error_message,
    )
    latest_job = db.get_latest_document_job(workflow_id)
    if latest_job:
        job_updates = {"current_stage": stage}
        if stage in {"ocr_review", "translation_review", "chunk_review", "ready_for_ingestion"}:
            job_updates["status"] = "waiting_review"
        elif stage == "completed":
            job_updates["status"] = "completed"
            job_updates["completed_at"] = datetime.utcnow().isoformat()
        elif stage == "failed":
            job_updates["status"] = "failed"
            job_updates["completed_at"] = datetime.utcnow().isoformat()
            job_updates["error_message"] = error_message
        else:
            job_updates["status"] = "running"
        db.update_document_job(latest_job["id"], **job_updates)
    return {"updated": True, "stage": stage}


@activity.defn
async def persist_document_content(workflow_id: str, pages: list[dict], chunks: list[dict]) -> dict:
    """Persist pages and chunks to SQLite."""
    from .. import db

    activity.logger.info(f"Persisting content for {workflow_id}: {len(pages)} pages, {len(chunks)} chunks")
    db.save_pages(workflow_id, pages)
    db.save_chunks(workflow_id, chunks)
    return {"persisted": True, "pages": len(pages), "chunks": len(chunks)}


@activity.defn
async def auto_tag_chunks_from_db(workflow_id: str, filename: str = "") -> dict:
    """Auto-assign domain tags to chunks using the configured LLM tagger."""
    from .. import db
    from ..domain_tags.base import load_taxonomy_for_instance, validate_tags_against_taxonomy
    from ..domain_tags.gemma_tagger import auto_tag_chunks
    from ..domain_tags.service import get_domain_tagger, load_domain_tagging_config

    config = load_domain_tagging_config()
    if not config.enabled:
        activity.logger.info("Domain tagging disabled; skipping workflow %s", workflow_id)
        return {"tagged_chunks": 0, "skipped": True}

    chunks = db.get_chunks(workflow_id, include_excluded=True)
    if not chunks:
        return {"tagged_chunks": 0, "skipped": True}

    doc = db.get_document(workflow_id) or {}
    doc_context_parts = [
        doc.get("source_manifest_name") or "",
        doc.get("display_name") or "",
    ]
    doc_context = " | ".join(part for part in doc_context_parts if part)
    taxonomy = load_taxonomy_for_instance(doc.get("instance"))

    tagger = get_domain_tagger(config)
    total_chunks = len(chunks)
    _activity_heartbeat(
        {
            "workflow_id": workflow_id,
            "stage": "auto_tag",
            "chunks_total": total_chunks,
            "chunks_completed": 0,
        }
    )

    def _tag_progress(event: dict) -> None:
        _activity_heartbeat(
            {
                "workflow_id": workflow_id,
                "stage": "auto_tag",
                "chunks_total": int(event.get("chunks_total") or total_chunks),
                "chunks_completed": int(event.get("chunks_completed") or 0),
                "chunk_number": int(event.get("chunk_number") or 0),
            }
        )

    tagged_map = await auto_tag_chunks(
        chunks,
        filename=filename or doc.get("filename") or "",
        doc_context=doc_context,
        tagger=tagger,
        taxonomy=taxonomy,
        log=activity.logger.info,
        progress_callback=_tag_progress,
    )

    db.delete_auto_chunk_tags(workflow_id)
    tagged_chunks = 0
    total_tags = 0
    for chunk_num, tags in tagged_map.items():
        if config.strict_taxonomy:
            tags = validate_tags_against_taxonomy(tags, taxonomy, strict=True)
        if not tags:
            continue
        db.replace_chunk_tags(
            workflow_id,
            chunk_num,
            [{"dimension": t.dimension, "value": t.value} for t in tags],
            source="auto",
        )
        tagged_chunks += 1
        total_tags += len(tags)

    activity.logger.info(
        "Auto domain tagging complete for %s: %s chunks, %s tags",
        workflow_id,
        tagged_chunks,
        total_tags,
    )
    return {"tagged_chunks": tagged_chunks, "total_tags": total_tags, "skipped": False}


@activity.defn
async def detect_and_translate_pages(
    pages: list[dict],
    target_language: str = "en",
    source_language: str = None,
    force_retranslate: bool = False,
) -> list[dict]:
    """Detect language and translate non-English pages."""
    return await _detect_and_translate_impl(
        pages,
        target_language=target_language,
        source_language=source_language,
        force_retranslate=force_retranslate,
    )


@activity.defn
async def detect_and_translate_pages_from_db(
    workflow_id: str,
    target_language: str = "en",
    source_language: str = None,
    force_retranslate: bool = False,
) -> dict:
    """Detect and translate pages loaded from SQLite; persist updated pages back to SQLite."""
    from .. import db

    pages = db.get_pages(workflow_id)
    from ..services import progress_cache
    from ..services.progress import publish_live_progress

    def _persist_translation_progress(*, phase: str, pages_completed: int, pages_total: int, **extra: object) -> dict:
        updated_at = datetime.utcnow().isoformat()
        mapped_phase = phase if phase in {"ocr", "translation", "chunking", "ingest"} else "translation"
        prev = progress_cache.get(workflow_id) or {}
        pages_completed, pages_total = _coalesce_translation_progress(
            event_phase=str(phase or "translation"),
            pages_completed=pages_completed,
            pages_total=pages_total,
            prev=prev,
            sqlite_translated=db.count_translated_pages(workflow_id),
            document_page_count=len(pages),
        )
        payload = {
            "workflow_id": workflow_id,
            "stage": "translation",
            "phase": mapped_phase,
            "pages_total": pages_total,
            "pages_completed": pages_completed,
            "done": pages_completed,
            "total": pages_total,
            "unit": "pages",
            "updated_at": updated_at,
            "force_retranslate": force_retranslate,
            **extra,
        }
        publish_live_progress(payload)
        return payload

    _activity_heartbeat(_persist_translation_progress(phase="translation", pages_completed=0, pages_total=len(pages)))

    def _translation_progress(event: dict) -> None:
        phase = str(event.get("phase") or "translation")
        if phase == "translation" and event.get("translated_page"):
            # Persist page-level wins immediately so retries skip already completed work.
            db.save_pages(workflow_id, [event["translated_page"]])
        payload = _persist_translation_progress(
            phase=phase,
            pages_completed=int(event.get("pages_completed") or 0),
            pages_total=len(pages),
            translated_count=int(event.get("translated_count") or 0),
            failed_count=int(event.get("failed_count") or 0),
        )
        _activity_heartbeat(payload)

    translated = await _detect_and_translate_impl(
        pages,
        target_language=target_language,
        source_language=source_language,
        force_retranslate=force_retranslate,
        progress_callback=_translation_progress,
    )
    db.save_pages(workflow_id, translated)
    translated_count = sum(1 for p in translated if p.get("translated_markdown"))
    latest_job = db.get_latest_document_job(workflow_id)
    translation_config = load_translation_config(target_language=target_language)
    translation_instance = (db.get_document(workflow_id) or {}).get("instance")
    translated_json_path = _write_json_temp(translated)
    try:
        translated_uri, translated_size, translated_mime = _upload_file_to_minio(
            translated_json_path, workflow_id, "translation_pages_json", "translated_pages.json",
            instance=translation_instance,
        )
    finally:
        if os.path.exists(translated_json_path):
            os.remove(translated_json_path)
    db.add_document_artifact(
        workflow_id=workflow_id,
        job_id=latest_job["id"] if latest_job else None,
        artifact_type="translation_pages_json",
        stage="translation_review",
        storage_uri=translated_uri,
        mime_type=translated_mime,
        filename="translated_pages.json",
        size_bytes=translated_size,
        metadata={
            "page_count": len(translated),
            "translated_count": translated_count,
            "translation_provider": translation_config.provider,
            "translation_model": translation_config.model,
            "translation_target_language": target_language,
            "translation_run_id": str(uuid4()),
        },
    )
    return {"page_count": len(translated), "translated_count": translated_count}
