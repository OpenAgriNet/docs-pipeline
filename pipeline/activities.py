"""
Temporal activities for the OCR pipeline.
Each activity is a retryable unit of work.
"""

import asyncio
import base64
import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
import tempfile
import mimetypes
from datetime import datetime
from pathlib import Path
from threading import Lock
from uuid import uuid4

import httpx
import fitz
import tiktoken
from langchain_text_splitters import RecursiveCharacterTextSplitter
from minio import Minio
from pypdf import PdfReader, PdfWriter
from temporalio import activity

from .chunking import chunk_pages, load_chunking_config
from .ocr import ocr_pdf as run_ocr_pdf, ocr_pdf_in_segments as run_ocr_pdf_in_segments
from .translation import load_translation_config, translate_pages

SUPPORTED_INPUT_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".ppt", ".pptx", ".xls", ".xlsx", ".csv", ".jpg", ".jpeg", ".png", ".webp", ".tif", ".tiff"
}
IMAGE_INPUT_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".tif", ".tiff"}
OFFICE_INPUT_EXTENSIONS = {".doc", ".docx", ".ppt", ".pptx", ".xls", ".xlsx"}
DELIMITED_INPUT_EXTENSIONS = {".csv"}
NATIVE_SPREADSHEET_EXTENSIONS = {".xlsx"}

_doc_metadata_cache: dict[str, dict] = {}
_doc_descriptions_cache: dict[str, str] = {}
_metadata_loaded = False
_metadata_lock = Lock()


def get_minio_client():
    """Get MinIO client from environment. Credentials are required."""
    access_key = os.environ.get("MINIO_ACCESS_KEY")
    secret_key = os.environ.get("MINIO_SECRET_KEY")

    if not access_key or not secret_key:
        raise RuntimeError("MINIO_ACCESS_KEY and MINIO_SECRET_KEY environment variables are required")

    return Minio(
        os.environ.get("MINIO_ENDPOINT", "localhost:9000"),
        access_key=access_key,
        secret_key=secret_key,
        secure=False,
    )


def download_from_minio(minio_path: str) -> str:
    """Download file from MinIO and return local temp path."""
    path = minio_path.replace("minio://", "")
    parts = path.split("/", 1)
    bucket = parts[0]
    object_name = parts[1] if len(parts) > 1 else ""

    client = get_minio_client()

    suffix = Path(object_name).suffix
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    temp_path = temp_file.name
    temp_file.close()

    client.fget_object(bucket, object_name, temp_path)
    return temp_path


def _minio_object_name(instance: str | None, workflow_id: str, artifact_type: str, filename: str) -> str:
    """Object key for a document artifact, prefixed by its tenant for isolation.

    Layout: ``<instance>/<workflow_id>/<artifact_type>/<filename>``. New writes
    always carry the tenant prefix (including the default tenant). Reads never
    reconstruct this key — they use the stored ``minio://`` URI on the artifact
    row — so pre-prefix objects remain readable.
    """
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", filename or "artifact")
    inst = _normalize_instance(instance)
    return f"{inst}/{workflow_id}/{artifact_type}/{safe_name}"


def _upload_file_to_minio(
    local_path: str,
    workflow_id: str,
    artifact_type: str,
    filename: str,
    instance: str | None = None,
) -> tuple[str, int, str]:
    client = get_minio_client()
    bucket = os.environ.get("MINIO_BUCKET", "documents")
    object_name = _minio_object_name(instance, workflow_id, artifact_type, filename)
    mime_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    client.fput_object(bucket, object_name, local_path, content_type=mime_type)
    size_bytes = os.path.getsize(local_path)
    return (f"minio://{bucket}/{object_name}", size_bytes, mime_type)


def _write_json_temp(data: object, suffix: str = ".json") -> str:
    with tempfile.NamedTemporaryFile("w", delete=False, suffix=suffix, encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        return f.name


def _source_type_from_path(filepath: str) -> tuple[str, str]:
    ext = Path(filepath).suffix.lower()
    if ext in {".csv", ".xlsx"}:
        return ("spreadsheet", "spreadsheet")
    if ext in IMAGE_INPUT_EXTENSIONS | OFFICE_INPUT_EXTENSIONS | {".pdf"}:
        return ("document", "pdf")
    return ("unknown", "pdf")


def _normalized_filename(original_name: str, canonical_input_type: str) -> str:
    stem = Path(original_name).stem or "document"
    if canonical_input_type == "spreadsheet":
        ext = Path(original_name).suffix.lower()
        return f"{stem}{ext if ext in {'.csv', '.xlsx'} else '.csv'}"
    return f"{stem}.pdf"


def _resolve_local_path(filepath: str) -> tuple[str, bool]:
    """
    Resolve path into local file path.
    Returns (path, should_cleanup).
    """
    if filepath.startswith("minio://") or filepath.startswith("minio:/"):
        minio_path = filepath
        if filepath.startswith("minio:/") and not filepath.startswith("minio://"):
            minio_path = filepath.replace("minio:/", "minio://", 1)
        local_path = download_from_minio(minio_path)
        activity.logger.info(f"Downloaded from MinIO to {local_path}")
        return local_path, True
    return filepath, False


def _normalize_filename(name: str) -> str:
    return (name or "").strip().lower()


def _load_metadata_once() -> None:
    global _metadata_loaded
    if _metadata_loaded:
        return
    with _metadata_lock:
        if _metadata_loaded:
            return

        metadata_csv_path = os.getenv(
            "DOCUMENT_METADATA_CSV_PATH", "/app/workspace/document_manifest.csv"
        )
        descriptions_jsonl_path = os.getenv(
            "DOCUMENT_DESCRIPTIONS_JSONL_PATH",
            "/app/workspace/document_descriptions.jsonl",
        )

        if os.path.exists(metadata_csv_path):
            try:
                with open(metadata_csv_path, "r", encoding="utf-8-sig", newline="") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        file_name = (row.get("File Name") or "").strip()
                        if not file_name or file_name.startswith("http://") or file_name.startswith("https://"):
                            continue
                        key = _normalize_filename(file_name)
                        _doc_metadata_cache[key] = {
                            "title_en": (row.get("Title (English)") or "").strip(),
                            "title_gu": (row.get("Title (Gujarati)") or "").strip(),
                            "doc_language": (row.get("Language (Gujarati / English)") or "").strip(),
                            "category_tags": (row.get("Category Tags (") or "").strip(),
                            "doc_short_description": (row.get("Description") or "").strip(),
                            "quality_score": (row.get("Quality(1-5)") or "").strip(),
                            "priority_rank": (row.get("Priority(1-5)") or "").strip(),
                            "ingestion_status": (row.get("Status ingested in the system") or "").strip(),
                        }
            except Exception as e:
                activity.logger.warning(f"Failed loading document metadata CSV: {e}")

        if os.path.exists(descriptions_jsonl_path):
            try:
                with open(descriptions_jsonl_path, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if not line:
                            continue
                        obj = json.loads(line)
                        key = _normalize_filename(obj.get("file", ""))
                        if key and obj.get("description"):
                            _doc_descriptions_cache[key] = str(obj["description"]).strip()
            except Exception as e:
                activity.logger.warning(f"Failed loading document descriptions JSONL: {e}")

        _metadata_loaded = True


def _get_doc_metadata(filename: str) -> dict:
    _load_metadata_once()
    return _doc_metadata_cache.get(_normalize_filename(filename), {})


def _get_doc_description(filename: str) -> str:
    _load_metadata_once()
    return _doc_descriptions_cache.get(_normalize_filename(filename), "")


def _convert_image_to_pdf(input_path: str, output_path: str) -> None:
    from PIL import Image, ImageOps

    with Image.open(input_path) as img:
        rgb_img = ImageOps.exif_transpose(img).convert("RGB")
        rgb_img.save(output_path, "PDF", resolution=300.0)


def _convert_office_to_pdf(input_path: str, output_dir: str) -> str:
    soffice_bin = shutil.which("soffice")
    if not soffice_bin:
        raise RuntimeError(
            "LibreOffice (soffice) is not installed. Install libreoffice to convert office files to PDF."
        )

    cmd = [
        soffice_bin,
        "--headless",
        "--nologo",
        "--nodefault",
        "--nolockcheck",
        "--nofirststartwizard",
        "--convert-to",
        "pdf",
        "--outdir",
        output_dir,
        input_path,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(
            f"Failed to convert office document to PDF: {result.stderr.strip() or result.stdout.strip()}"
        )

    output_pdf = Path(output_dir) / f"{Path(input_path).stem}.pdf"
    if not output_pdf.exists():
        raise RuntimeError("Office-to-PDF conversion finished but output PDF was not found")
    return str(output_pdf)


def _ensure_pdf_input(local_path: str) -> tuple[str, bool]:
    """
    Ensure the given file path points to a PDF.
    Returns (pdf_path, should_cleanup).
    """
    ext = Path(local_path).suffix.lower()
    if ext not in SUPPORTED_INPUT_EXTENSIONS:
        raise ValueError(
            f"Unsupported file extension '{ext}'. Supported extensions: {sorted(SUPPORTED_INPUT_EXTENSIONS)}"
        )

    if ext == ".pdf":
        return local_path, False

    work_dir = tempfile.mkdtemp(prefix="doc_convert_")
    output_pdf = Path(work_dir) / f"{Path(local_path).stem}.pdf"

    if ext in IMAGE_INPUT_EXTENSIONS:
        _convert_image_to_pdf(local_path, str(output_pdf))
    elif ext in OFFICE_INPUT_EXTENSIONS:
        converted = _convert_office_to_pdf(local_path, work_dir)
        if converted != str(output_pdf):
            shutil.move(converted, output_pdf)
    else:
        raise ValueError(f"Unsupported conversion path for extension: {ext}")

    activity.logger.info(f"Converted {local_path} -> {output_pdf}")
    return str(output_pdf), True


def _csv_to_pages(input_path: str, rows_per_page: int = 80) -> list[dict]:
    import csv

    def _open_reader():
        last_err = None
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                f = open(input_path, "r", encoding=enc, newline="")
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except Exception:
                    dialect = csv.excel
                return f, csv.reader(f, dialect=dialect)
            except Exception as e:
                last_err = e
        raise RuntimeError(f"Could not read CSV file: {last_err}")

    f, reader = _open_reader()
    try:
        rows = [[(c or "").strip() for c in r] for r in reader if any((c or "").strip() for c in r)]
    finally:
        f.close()

    if not rows:
        return [{
            "page_number": 1,
            "original_markdown": f"# {Path(input_path).name}\n\n(Empty CSV file)",
            "edited_markdown": None,
            "is_reviewed": False,
            "reviewer_notes": None,
        }]

    header = rows[0]
    body = rows[1:] if len(rows) > 1 else []

    pages: list[dict] = []
    page_num = 1
    for i in range(0, len(body), rows_per_page):
        batch = body[i:i + rows_per_page]
        lines = [f"# Spreadsheet Data: {Path(input_path).name}", "", f"Columns: {', '.join(header)}", ""]
        for row_idx, row in enumerate(batch, start=i + 1):
            pairs = []
            for col_i, val in enumerate(row):
                col = header[col_i] if col_i < len(header) and header[col_i] else f"col_{col_i+1}"
                pairs.append(f"{col}: {val}")
            if pairs:
                lines.append(f"- Row {row_idx}: " + " | ".join(pairs))
        pages.append({
            "page_number": page_num,
            "original_markdown": "\n".join(lines),
            "edited_markdown": None,
            "is_reviewed": False,
            "reviewer_notes": None,
        })
        page_num += 1
    return pages


def _xlsx_to_pages(input_path: str, rows_per_page: int = 80) -> list[dict]:
    from datetime import date, datetime
    from openpyxl import load_workbook

    def _cell_to_str(value) -> str:
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        return str(value).strip()

    pages: list[dict] = []
    page_num = 1
    wb = load_workbook(filename=input_path, data_only=True, read_only=True)

    for sheet in wb.worksheets:
        header: list[str] | None = None
        batch: list[tuple[int, list[str]]] = []
        seen_rows = 0

        def emit_page(rows_batch: list[tuple[int, list[str]]]) -> None:
            nonlocal page_num
            if not rows_batch:
                return
            assert header is not None
            lines = [
                f"# Spreadsheet Data: {Path(input_path).name}",
                "",
                f"Sheet: {sheet.title}",
                f"Columns: {', '.join(header)}",
                "",
            ]
            for excel_row_num, row in rows_batch:
                pairs = []
                for col_i, val in enumerate(row):
                    col = header[col_i] if col_i < len(header) and header[col_i] else f"col_{col_i+1}"
                    pairs.append(f"{col}: {val}")
                if pairs:
                    lines.append(f"- Row {excel_row_num}: " + " | ".join(pairs))
            pages.append({
                "page_number": page_num,
                "original_markdown": "\n".join(lines),
                "edited_markdown": None,
                "is_reviewed": False,
                "reviewer_notes": None,
            })
            page_num += 1

        for excel_row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [_cell_to_str(v) for v in row]
            if not any(values):
                continue
            seen_rows += 1

            if header is None:
                header = values
                continue

            batch.append((excel_row_num, values))
            if len(batch) >= rows_per_page:
                emit_page(batch)
                batch = []

        if header is not None:
            emit_page(batch)
        elif seen_rows == 0:
            pages.append({
                "page_number": page_num,
                "original_markdown": f"# Spreadsheet Data: {Path(input_path).name}\n\nSheet: {sheet.title}\n\n(Empty sheet)",
                "edited_markdown": None,
                "is_reviewed": False,
                "reviewer_notes": None,
            })
            page_num += 1

    wb.close()

    if not pages:
        return [{
            "page_number": 1,
            "original_markdown": f"# Spreadsheet Data: {Path(input_path).name}\n\n(Empty workbook)",
            "edited_markdown": None,
            "is_reviewed": False,
            "reviewer_notes": None,
        }]
    return pages


# =============================================================================
# Text Processing Utilities
# =============================================================================

def count_tokens(text: str, model: str = "cl100k_base") -> int:
    if not text:
        return 0
    encoder = tiktoken.get_encoding(model)
    return len(encoder.encode(str(text), disallowed_special=()))


def clean_html_tags(text: str) -> str:
    if not isinstance(text, str):
        return text
    text = text.replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
    text = re.sub(r"<[^>]+>", "", text)
    return text.strip()


def clean_latex_notation(text: str) -> str:
    if not isinstance(text, str):
        return text
    text = re.sub(r"\[\^[0-9]+\]", "", text)
    text = re.sub(r"\$\s*\{\s*\}\s*\^\{[0-9]+\}\s*\$", "", text)
    text = re.sub(r"\$\s*\^\{[0-9]+\}\s*\$", "", text)
    text = re.sub(r"\$\s*\$", "", text)
    text = re.sub(r"\\[a-zA-Z]+\{[^}]*\}", "", text)
    text = re.sub(r"\\[a-zA-Z]+", "", text)
    text = re.sub(r"[\\{}]", "", text)
    return text


def format_table_content(text: str) -> str:
    if not isinstance(text, str):
        return text
    lines = text.split("\n")
    cleaned_lines = []
    for line in lines:
        if re.match(r"^[\s\|]*$", line):
            continue
        if re.match(r"^[\s\|\-\:]*$", line):
            continue
        line = re.sub(r"\|\s*\|", "|", line)
        line = re.sub(r"^\|\s*", "", line)
        line = re.sub(r"\s*\|$", "", line)
        line = re.sub(r"\s+", " ", line).strip()
        if line:
            cleaned_lines.append(line)
    return "\n".join(cleaned_lines)


def clean_text(text: str) -> str:
    if not isinstance(text, str):
        return text
    text = clean_html_tags(text)
    text = clean_latex_notation(text)
    text = format_table_content(text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    lines = [line.strip() for line in text.split("\n")]
    text = "\n".join(lines)
    text = re.sub(r"\n\s*\n\s*\n", "\n\n", text)
    text = re.sub(r"\n ", "\n", text)
    text = re.sub(r" \n", "\n", text)
    return text.strip()


def _infer_section(text: str, section_title: str | None = None) -> str:
    """Best-effort section heading for Marqo provenance."""
    if section_title and str(section_title).strip():
        return str(section_title).strip()
    if not text:
        return ""
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("## "):
            return stripped[3:].strip()
        if stripped.startswith("# "):
            return stripped[2:].strip()
    return ""


def is_reference_section(text: str) -> bool:
    """
    Detect if text is primarily a reference/bibliography section.
    Returns True if the text appears to be citations/references.
    """
    if not text or len(text) < 50:
        return False

    ref_headers = [
        r"^\s*#{1,3}\s*(?:references|bibliography|citations|works cited|literature cited)\s*$",
        r"^\s*\*{1,2}(?:references|bibliography)\*{1,2}\s*$",
    ]
    for pattern in ref_headers:
        if re.search(pattern, text, re.IGNORECASE | re.MULTILINE):
            return True

    lines = text.split("\n")
    total_lines = len([l for l in lines if l.strip()])
    if total_lines == 0:
        return False

    citation_patterns = [
        r"^\s*\d{1,3}[\.\)]\s+[A-Z][a-z]+[\s,].*(?:\d{4}|\(\d{4}\))",
        r"doi[:\s]*10\.\d{4,}",
        r"(?:J\.|Journal|Int\.|Proceedings|Trans\.).*\d{4}",
        r"\(\d{4}\)\s*$",
        r"\bet\s+al\b",
        r"(?:Vol\.?\s*\d+|\d+\s*\(\d+\)\s*:)",
        r"(?:pp?\.?\s*\d+[-–]\d+|:\s*\d+[-–]\d+)",
    ]

    citation_line_count = 0
    for line in lines:
        if not line.strip():
            continue
        for pattern in citation_patterns:
            if re.search(pattern, line, re.IGNORECASE):
                citation_line_count += 1
                break

    return (citation_line_count / total_lines) > 0.4


def _ocr_pdf(local_pdf_path: str) -> list[dict]:
    return run_ocr_pdf(local_pdf_path, clean_text)


def _ocr_pdf_in_segments(
    local_pdf_path: str,
    segment_pages: int,
    on_segment_complete=None,
    completed_page_numbers: set[int] | None = None,
) -> list[dict]:
    return run_ocr_pdf_in_segments(
        local_pdf_path,
        segment_pages,
        clean_text,
        on_segment_complete=on_segment_complete,
        completed_page_numbers=completed_page_numbers,
    )


def _build_chunks_from_pages(
    pages: list[dict],
    chunk_size: int = 450,
    chunk_overlap: int = 128,
    min_tokens: int = 100,
) -> list[dict]:
    raise RuntimeError("_build_chunks_from_pages is deprecated; use chunk_pages() via create_chunks_from_db")


def clean_text_for_ingestion(text: str) -> str:
    """Clean translation preambles from text before ingestion."""
    if not text:
        return text

    result = text
    result = re.sub(
        r"^Here is the translated text from \*\*[^*]+\*\* to English[^:]*:?\s*\n*-{0,3}\s*\n*",
        "",
        result,
        flags=re.IGNORECASE,
    )
    result = re.sub(
        r"^Here is the translated text from [^:]+?:\s*\n*-{0,3}\s*\n*",
        "",
        result,
        flags=re.IGNORECASE,
    )
    result = re.sub(
        r"^Here is the translated text[^:]*:?\s*\n*-{0,3}\s*\n*",
        "",
        result,
        flags=re.IGNORECASE,
    )
    result = re.sub(
        r"^Here is the (?:English )?translation[^:]*:?\s*\n*",
        "",
        result,
        flags=re.IGNORECASE | re.MULTILINE,
    )
    result = re.sub(
        r"^Here is the translated text with[^:]+:?\s*\n*-{0,3}\s*\n*",
        "",
        result,
        flags=re.IGNORECASE,
    )

    prefixes = [
        r"^(?:the\s+)?english\s+translation:?\s*\n*",
        r"^(?:the\s+)?translation:?\s*\n*",
        r"^translated\s+(?:text|content):?\s*\n*",
        r"^##?\s*(?:english\s+)?translation\s*\n+",
        r"^---+\s*\n+",
        r"^\*\*Translation:?\*\*\s*\n*",
    ]
    for pattern in prefixes:
        result = re.sub(pattern, "", result, flags=re.IGNORECASE | re.MULTILINE)

    result = re.sub(r"\n*-{3,}\s*$", "", result)
    return result.strip()


def _normalize_instance(value: str | None) -> str:
    """Ingest-side instance normalizer (mirrors auth.tenancy without importing FastAPI)."""
    text = (value or "").strip().lower()
    if text:
        return text
    return (os.environ.get("DEFAULT_INSTANCE") or "default").strip().lower() or "default"


def _prepare_records(
    document_id: str,
    filename: str,
    chunks: list[dict],
    workflow_id: str | None = None,
    name_gu: str | None = None,
    name_en: str | None = None,
    description: str | None = None,
    include_e5_prefix_field: bool = True,
    instance: str | None = None,
) -> list[dict]:
    metadata = _get_doc_metadata(filename)
    resolved_instance = _normalize_instance(instance)
    llm_doc_description = _get_doc_description(filename)

    doc_hash = hashlib.md5(document_id.encode()).hexdigest()
    external_slug = workflow_id or filename
    default_name = filename.replace(".pdf", "").replace(".PDF", "")
    name_gu = name_gu or metadata.get("title_gu") or default_name
    name_en = name_en or metadata.get("title_en") or default_name
    category_tags = metadata.get("category_tags", "")
    quality_score = metadata.get("quality_score", "")
    priority_rank = metadata.get("priority_rank", "")
    ingestion_status = metadata.get("ingestion_status", "")
    doc_language = metadata.get("doc_language", "")
    short_description = metadata.get("doc_short_description", "")
    effective_description = description or llm_doc_description or short_description

    records = []
    for chunk in chunks:
        if chunk.get("is_excluded", False):
            continue

        raw_text = chunk.get("edited_text") or chunk.get("original_text", "")
        chunk_num = chunk.get("chunk_number", 0)
        text = clean_text_for_ingestion(raw_text)
        is_ref = is_reference_section(text)

        section = _infer_section(text, chunk.get("section_title"))
        record = {
            "_id": hashlib.md5(f"{doc_hash}_{chunk_num}_{text[:50]}".encode()).hexdigest(),
            "doc_id": document_id,
            "workflow_id": workflow_id or "",
            "instance": resolved_instance,
            "type": "document",
            "source": "docs-pipeline",
            "filename": external_slug,
            "name_gu": name_gu,
            "name_en": name_en,
            "title_en": metadata.get("title_en", ""),
            "title_gu": metadata.get("title_gu", ""),
            "doc_language": doc_language,
            "category_tags": category_tags,
            "doc_short_description": short_description,
            "doc_llm_description": llm_doc_description,
            "ingestion_status": ingestion_status,
            "description": effective_description,
            "text": text,
            "chunk_num": chunk_num,
            "section": section,
            "token_count": chunk.get("token_count", 0),
            "page_start": chunk.get("page_start", 1),
            "page_end": chunk.get("page_end", 1),
            "is_reference": is_ref,
            "quality_score": float(quality_score) if str(quality_score).strip().replace(".", "", 1).isdigit() else 0.0,
            "priority_rank": float(priority_rank) if str(priority_rank).strip().replace(".", "", 1).isdigit() else 0.0,
        }
        if include_e5_prefix_field:
            record["text_for_embedding"] = f"passage: {text}" if text else "passage:"
        domain_tags_flat = (chunk.get("domain_tags_flat") or "").strip()
        if domain_tags_flat:
            from .domain_tags.base import normalize_marqo_domain_tags_field

            record["domain_tags"] = normalize_marqo_domain_tags_field(domain_tags_flat)
        records.append(record)

    return records


def prepare_ingestion_records(
    document_id: str,
    filename: str,
    chunks: list[dict],
    workflow_id: str | None = None,
    name_gu: str | None = None,
    name_en: str | None = None,
    description: str | None = None,
    instance: str | None = None,
) -> list[dict]:
    """Public helper used by tests and scripts when preparing Marqo payloads."""
    return _prepare_records(
        document_id,
        filename,
        chunks,
        workflow_id=workflow_id,
        name_gu=name_gu,
        name_en=name_en,
        description=description,
        instance=instance,
    )


def _passage_schema_field_names() -> set[str]:
    """Field names for the canonical passage schema (E5 text_for_embedding + full metadata)."""
    settings = _marqo_settings(use_tensor_prefix_field=True)
    return {f.get("name") for f in settings.get("allFields", []) if isinstance(f, dict) and f.get("name")}


def _core_passage_schema_field_names() -> set[str]:
    """Required Marqo fields; optional fields like domain_tags and instance do not force index recreation."""
    return _passage_schema_field_names() - {"domain_tags", "instance"}


def _marqo_settings(use_tensor_prefix_field: bool = True) -> dict:
    tensor_field = "text_for_embedding" if use_tensor_prefix_field else "text"
    all_fields = [
        {"name": "doc_id", "type": "text", "features": ["filter"]},
        {"name": "workflow_id", "type": "text", "features": ["filter"]},
        {"name": "instance", "type": "text", "features": ["filter"]},
        {"name": "type", "type": "text", "features": ["filter"]},
        {"name": "source", "type": "text", "features": ["filter"]},
        {"name": "filename", "type": "text", "features": ["filter"]},
        {"name": "name_gu", "type": "text", "features": ["filter"]},
        {"name": "name_en", "type": "text", "features": ["filter"]},
        {"name": "title_en", "type": "text", "features": ["filter"]},
        {"name": "title_gu", "type": "text", "features": ["filter"]},
        {"name": "doc_language", "type": "text", "features": ["filter"]},
        {"name": "category_tags", "type": "text", "features": ["filter"]},
        {"name": "doc_short_description", "type": "text", "features": ["filter"]},
        {"name": "doc_llm_description", "type": "text", "features": ["filter"]},
        {"name": "ingestion_status", "type": "text", "features": ["filter"]},
        {"name": "description", "type": "text", "features": ["lexical_search"]},
        {"name": "chunk_num", "type": "int", "features": ["filter"]},
        {"name": "section", "type": "text", "features": ["filter"]},
        {"name": "token_count", "type": "int", "features": ["filter"]},
        {"name": "page_start", "type": "int", "features": ["filter"]},
        {"name": "page_end", "type": "int", "features": ["filter"]},
        {"name": "is_reference", "type": "bool", "features": ["filter"]},
        {"name": "quality_score", "type": "float", "features": ["filter"]},
        {"name": "priority_rank", "type": "float", "features": ["filter"]},
        {"name": "domain_tags", "type": "text", "features": ["filter"]},
        {"name": "text", "type": "text", "features": ["lexical_search"]},
        {"name": "priority", "type": "float", "features": ["score_modifier", "filter"]},
    ]
    if use_tensor_prefix_field:
        all_fields.append({"name": "text_for_embedding", "type": "text"})

    return {
        "type": "structured",
        "vectorNumericType": "float",
        "model": "hf/multilingual-e5-large",
        "normalizeEmbeddings": False,
        "textPreprocessing": {"splitLength": 3, "splitOverlap": 1, "splitMethod": "sentence"},
        "allFields": all_fields,
        "tensorFields": [tensor_field],
    }


async def _detect_and_translate_impl(
    pages: list[dict],
    target_language: str = "en",
    source_language: str | None = None,
) -> list[dict]:
    del source_language
    config = load_translation_config(target_language=target_language)
    return await translate_pages(
        pages,
        target_language=target_language,
        config=config,
        log=activity.logger.info,
    )


# =============================================================================
# Activities
# =============================================================================

@activity.defn
async def run_ocr(filepath: str) -> list[dict]:
    """Run OCR on a supported file and return page dicts."""
    activity.logger.info(f"Running OCR on {filepath}")

    local_path, cleanup_local = _resolve_local_path(filepath)
    ext = Path(local_path).suffix.lower()
    pdf_path = local_path
    cleanup_pdf_dir = False

    try:
        if ext in DELIMITED_INPUT_EXTENSIONS:
            pages = _csv_to_pages(local_path)
            activity.logger.info(f"CSV parsed into {len(pages)} pages")
            return pages
        if ext in NATIVE_SPREADSHEET_EXTENSIONS:
            pages = _xlsx_to_pages(local_path)
            activity.logger.info(f"XLSX parsed into {len(pages)} pages")
            return pages
        pdf_path, cleanup_pdf_dir = _ensure_pdf_input(local_path)
        return _ocr_pdf(pdf_path)
    finally:
        if cleanup_pdf_dir:
            try:
                shutil.rmtree(Path(pdf_path).parent, ignore_errors=True)
            except Exception:
                pass
        if cleanup_local and os.path.exists(local_path):
            os.remove(local_path)


@activity.defn
async def run_ocr_and_store(workflow_id: str, filepath: str) -> dict:
    """Run OCR and persist pages to SQLite to avoid large Temporal payloads."""
    from . import db

    local_path, cleanup_local = _resolve_local_path(filepath)
    ext = Path(local_path).suffix.lower()
    source_type, canonical_input_type = _source_type_from_path(local_path)
    original_filename = Path(local_path).name
    normalized_path = local_path
    cleanup_normalized = False
    segment_pages = max(
        1,
        int(os.environ.get("OCR_SEGMENT_PAGES", "20")),
    )

    try:
        if ext in DELIMITED_INPUT_EXTENSIONS:
            pages = _csv_to_pages(local_path)
        elif ext in NATIVE_SPREADSHEET_EXTENSIONS:
            pages = _xlsx_to_pages(local_path)
        else:
            normalized_path, cleanup_normalized = _ensure_pdf_input(local_path)
            saved_page_numbers = set(db.get_saved_page_numbers(workflow_id))

            def persist_segment(segment_pages_result: list[dict], total_pages: int) -> None:
                db.save_pages(workflow_id, segment_pages_result)
                current_saved = len(saved_page_numbers.union({p["page_number"] for p in segment_pages_result}))
                saved_page_numbers.update(p["page_number"] for p in segment_pages_result)
                db.update_document_fields(workflow_id, page_count=current_saved)
                activity.heartbeat({"workflow_id": workflow_id, "pages_saved": current_saved, "total_pages": total_pages})
                activity.logger.info(
                    "Persisted OCR segment for %s: %s/%s pages saved",
                    workflow_id,
                    current_saved,
                    total_pages,
                )

            pages = await asyncio.to_thread(
                _ocr_pdf_in_segments,
                normalized_path,
                segment_pages=segment_pages,
                on_segment_complete=persist_segment,
                completed_page_numbers=saved_page_numbers,
            )

            pages = db.get_pages(workflow_id)

        db.save_pages(workflow_id, pages)

        latest_job = db.get_latest_document_job(workflow_id)
        job_id = latest_job["id"] if latest_job else None
        # Tenant prefix for new artifact writes (from the durable SQLite row).
        doc_instance = (db.get_document(workflow_id) or {}).get("instance")

        normalized_filename = _normalized_filename(original_filename, canonical_input_type)
        normalized_uri, normalized_size, normalized_mime = _upload_file_to_minio(
            normalized_path,
            workflow_id,
            "normalized_spreadsheet" if canonical_input_type == "spreadsheet" else "normalized_pdf",
            normalized_filename,
            instance=doc_instance,
        )
        normalized_artifact_id = db.add_document_artifact(
            workflow_id=workflow_id,
            job_id=job_id,
            artifact_type="normalized_spreadsheet" if canonical_input_type == "spreadsheet" else "normalized_pdf",
            stage="ocr_processing",
            storage_uri=normalized_uri,
            mime_type=normalized_mime,
            filename=normalized_filename,
            size_bytes=normalized_size,
            metadata={"source_filepath": filepath, "canonical_input_type": canonical_input_type},
        )

        if filepath.startswith("minio://"):
            original_uri = filepath
            original_size = None
            original_mime = mimetypes.guess_type(original_filename)[0] or "application/octet-stream"
        else:
            original_uri, original_size, original_mime = _upload_file_to_minio(
                local_path,
                workflow_id,
                "original_upload",
                original_filename,
                instance=doc_instance,
            )
        original_artifact_id = db.add_document_artifact(
            workflow_id=workflow_id,
            job_id=job_id,
            artifact_type="original_upload",
            stage="registered",
            storage_uri=original_uri,
            mime_type=original_mime,
            filename=original_filename,
            size_bytes=original_size,
            metadata={"source_filepath": filepath},
        )

        pages_json_path = _write_json_temp(pages)
        try:
            pages_uri, pages_size, pages_mime = _upload_file_to_minio(
                pages_json_path, workflow_id, "ocr_pages_json", "pages.json", instance=doc_instance
            )
        finally:
            if os.path.exists(pages_json_path):
                os.remove(pages_json_path)

        db.add_document_artifact(
            workflow_id=workflow_id,
            job_id=job_id,
            artifact_type="ocr_pages_json",
            stage="ocr_review",
            storage_uri=pages_uri,
            mime_type=pages_mime,
            filename="pages.json",
            size_bytes=pages_size,
            metadata={"page_count": len(pages)},
        )

        db.update_document_fields(
            workflow_id,
            page_count=len(pages),
            source_type=source_type,
            canonical_input_type=canonical_input_type,
            original_artifact_id=original_artifact_id,
            normalized_artifact_id=normalized_artifact_id,
        )
        return {"page_count": len(pages), "normalized_artifact_id": normalized_artifact_id}
    finally:
        if cleanup_normalized and os.path.exists(normalized_path):
            shutil.rmtree(Path(normalized_path).parent, ignore_errors=True)
        if cleanup_local and os.path.exists(local_path):
            os.remove(local_path)


@activity.defn
async def create_chunks(
    pages: list[dict],
    chunk_size: int = 450,
    chunk_overlap: int = 128,
    min_tokens: int = 100,
) -> list[dict]:
    """Create chunks from pages with page range tracking."""
    activity.logger.info(f"Creating chunks from {len(pages)} pages")
    config = load_chunking_config(chunk_size=chunk_size, chunk_overlap=chunk_overlap, min_tokens=min_tokens)
    result = await chunk_pages(pages, config)
    chunks = []
    for idx, chunk in enumerate(result.chunks, 1):
        chunks.append(
            {
                "chunk_number": idx,
                "original_text": chunk.text,
                "edited_text": None,
                "token_count": chunk.token_count,
                "page_start": chunk.page_start,
                "page_end": chunk.page_end,
                "source_page_numbers_json": json.dumps(chunk.source_page_numbers),
                "source_spans_json": json.dumps(chunk.source_spans),
                "section_title": chunk.section_title,
                "content_type": chunk.content_type,
                "is_reference": chunk.is_reference,
                "chunking_provider": result.provider,
                "chunking_model": result.model,
                "chunking_config_json": result.config.to_json(),
                "chunking_run_id": "",
                "chunk_version": 1,
                "is_reviewed": False,
                "is_excluded": False,
                "reviewer_notes": None,
            }
        )
    activity.logger.info(f"Created {len(chunks)} chunks with page tracking")
    return chunks


@activity.defn
async def create_chunks_from_db(
    workflow_id: str,
    chunk_size: int = 450,
    chunk_overlap: int = 128,
    min_tokens: int = 100,
) -> dict:
    """Create chunks from persisted pages and persist chunks in SQLite."""
    from . import db

    pages = db.get_pages(workflow_id)
    activity.logger.info(f"Creating chunks from DB pages for {workflow_id}: {len(pages)} pages")
    config = load_chunking_config(chunk_size=chunk_size, chunk_overlap=chunk_overlap, min_tokens=min_tokens)
    latest_job = db.get_latest_document_job(workflow_id)
    base_job_config = {}
    if latest_job and latest_job.get("config_json"):
        try:
            base_job_config = json.loads(latest_job["config_json"]) or {}
        except Exception:
            base_job_config = {}

    async def _persist_chunking_progress(event: dict) -> None:
        if not latest_job:
            return
        pages_total = int(event.get("pages_total") or len(pages) or 0)
        pages_processed = int(event.get("pages_processed") or 0)
        chunks_emitted = int(event.get("chunks_emitted") or 0)
        raw_percent = float(event.get("percent") or 0.0)
        percent = max(0.0, min(100.0, raw_percent))
        progress = {
            "status": "running" if percent < 100.0 else "completed",
            "provider": event.get("provider") or config.provider,
            "pages_processed": pages_processed,
            "pages_total": pages_total,
            "chunks_emitted": chunks_emitted,
            "percent": round(percent, 2),
            "updated_at": datetime.utcnow().isoformat(),
        }
        next_config = {**base_job_config, "chunking_progress": progress}
        db.update_document_job(latest_job["id"], config_json=next_config)

    await _persist_chunking_progress(
        {
            "provider": config.provider,
            "pages_processed": 0,
            "pages_total": len(pages),
            "chunks_emitted": 0,
            "percent": 0.0,
        }
    )

    result = await chunk_pages(pages, config, progress_callback=_persist_chunking_progress)
    chunking_run_id = f"chunk-{workflow_id}-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    chunks = []
    for idx, chunk in enumerate(result.chunks, 1):
        chunks.append(
            {
                "chunk_number": idx,
                "original_text": chunk.text,
                "edited_text": None,
                "token_count": chunk.token_count,
                "page_start": chunk.page_start,
                "page_end": chunk.page_end,
                "source_page_numbers_json": json.dumps(chunk.source_page_numbers),
                "source_spans_json": json.dumps(chunk.source_spans),
                "section_title": chunk.section_title,
                "content_type": chunk.content_type,
                "is_reference": chunk.is_reference,
                "chunking_provider": result.provider,
                "chunking_model": result.model,
                "chunking_config_json": result.config.to_json(),
                "chunking_run_id": chunking_run_id,
                "is_reviewed": False,
                "is_excluded": False,
                "reviewer_notes": None,
            }
        )
    db.save_chunks(workflow_id, chunks)
    chunks_instance = (db.get_document(workflow_id) or {}).get("instance")
    chunks_json_path = _write_json_temp(chunks)
    try:
        chunks_uri, chunks_size, chunks_mime = _upload_file_to_minio(
            chunks_json_path, workflow_id, "chunk_json_export", "chunks.json", instance=chunks_instance
        )
    finally:
        if os.path.exists(chunks_json_path):
            os.remove(chunks_json_path)
    db.add_document_artifact(
        workflow_id=workflow_id,
        job_id=latest_job["id"] if latest_job else None,
        artifact_type="chunk_json_export",
        stage="chunk_review",
        storage_uri=chunks_uri,
        mime_type=chunks_mime,
        filename="chunks.json",
        size_bytes=chunks_size,
        metadata={
            "chunk_count": len(chunks),
            "chunking_provider": result.provider,
            "chunking_model": result.model,
            "chunking_run_id": chunking_run_id,
            "chunking_config": json.loads(result.config.to_json()),
            "warnings": result.warnings,
            "stats": result.stats,
        },
    )
    # Reconcile the document row immediately after chunk persistence so SQLite
    # remains truthful even if the workflow fails before its final state update.
    db.reconcile_materialized_state(workflow_id)
    await _persist_chunking_progress(
        {
            "provider": result.provider,
            "pages_processed": len(pages),
            "pages_total": len(pages),
            "chunks_emitted": len(chunks),
            "percent": 100.0,
        }
    )
    return {"chunk_count": len(chunks)}


@activity.defn
async def prepare_for_ingestion(
    document_id: str,
    filename: str,
    chunks: list[dict],
    workflow_id: str | None = None,
    name_gu: str = None,
    name_en: str = None,
    description: str = None,
) -> list[dict]:
    """Prepare chunks for Marqo ingestion."""
    from . import db

    activity.logger.info(f"Preparing {len(chunks)} chunks for ingestion")
    doc = db.get_document(workflow_id) if workflow_id else None
    records = _prepare_records(
        document_id,
        filename,
        chunks,
        workflow_id=workflow_id,
        name_gu=name_gu,
        name_en=name_en,
        description=description,
        instance=(doc or {}).get("instance"),
    )
    activity.logger.info(f"Prepared {len(records)} records")
    return records


@activity.defn
async def ingest_to_marqo(
    records: list[dict],
    marqo_url: str = None,
    index_name: str = "documents-index",
    batch_size: int = 10,
) -> dict:
    """Ingest records to Marqo."""
    import marqo

    if not marqo_url:
        marqo_url = os.environ.get("MARQO_URL", "http://localhost:8882")

    activity.logger.info(f"Ingesting {len(records)} records to Marqo at {marqo_url}")
    mq = marqo.Client(url=marqo_url)

    settings = _marqo_settings(use_tensor_prefix_field=True)
    passage_fields = _passage_schema_field_names()
    core_passage_fields = _core_passage_schema_field_names()

    index_exists = True
    try:
        mq.get_index(index_name)
    except Exception:
        index_exists = False

    if not index_exists:
        mq.create_index(index_name, settings_dict=settings)
        activity.logger.info(f"Created index: {index_name} (passage schema)")
    else:
        index = mq.index(index_name)
        try:
            index_settings = index.get_settings()
            tensor_fields = set(index_settings.get("tensorFields", [])) if isinstance(index_settings, dict) else set()
            index_field_names = {
                f.get("name") for f in (index_settings.get("allFields") or [])
                if isinstance(f, dict) and f.get("name")
            }
            has_passage_tensor = "text_for_embedding" in tensor_fields
            has_full_schema = core_passage_fields <= index_field_names
            if not (has_passage_tensor and has_full_schema):
                mq.delete_index(index_name)
                mq.create_index(index_name, settings_dict=settings)
                activity.logger.info(
                    f"Recreated index: {index_name} with passage schema (was missing text_for_embedding or fields)"
                )
        except Exception as e:
            activity.logger.warning("Could not verify index schema, recreating: %s", e)
            try:
                mq.delete_index(index_name)
            except Exception:
                pass
            mq.create_index(index_name, settings_dict=settings)
            activity.logger.info(f"Recreated index: {index_name} (passage schema)")

    index = mq.index(index_name)
    allowed_fields = passage_fields
    if allowed_fields:
        try:
            index_field_names = {
                f.get("name") for f in (index.get_settings().get("allFields") or [])
                if isinstance(f, dict) and f.get("name")
            }
        except Exception:
            index_field_names = set()
        for i, record in enumerate(records):
            normalized = {"_id": record.get("_id")}
            for key, value in record.items():
                if key == "_id":
                    continue
                if key in allowed_fields:
                    # Optional fields absent from a legacy index would be rejected;
                    # skip them so the existing index needs no migration.
                    if key in ("domain_tags", "instance") and key not in index_field_names:
                        continue
                    normalized[key] = value
            records[i] = normalized

    for i in range(0, len(records), batch_size):
        batch = records[i : i + batch_size]
        result = index.add_documents(batch)
        if result.get("errors"):
            errors = []
            for item in result.get("items") or []:
                if item.get("status") != 200:
                    errors.append({
                        "_id": item.get("_id"),
                        "status": item.get("status"),
                        "error": item.get("error"),
                        "message": item.get("message"),
                        "code": item.get("code"),
                    })
            activity.logger.error(
                "Marqo add_documents reported errors. First few: %s. Full result keys: %s",
                errors[:5],
                list(result.keys()),
            )
            if errors:
                raise RuntimeError(
                    f"Marqo add_documents failed for {len(errors)} doc(s). First error: {errors[0]}"
                )

    stats = index.get_stats()
    activity.logger.info(f"Ingestion complete: {stats}")

    return {
        "records_ingested": len(records),
        "index_stats": stats,
        "supports_prefixed_tensor_field": True,
    }


@activity.defn
async def ingest_document_from_db(
    workflow_id: str,
    document_id: str,
    filename: str,
    marqo_url: str = None,
    index_name: str = "documents-index",
    batch_size: int = 10,
) -> dict:
    """Prepare and ingest chunks directly from SQLite by workflow_id."""
    from . import db

    chunks = db.get_chunks(workflow_id, include_excluded=True)
    doc = db.get_document(workflow_id)
    # Write chunks to the physical Marqo index that the document's tenant owns for
    # its (logical) index. Registry-resolved; falls back to the caller-supplied
    # index_name for the legacy single-index deployment (empty registry).
    resolved_index = db.resolve_marqo_index((doc or {}).get("instance"), (doc or {}).get("index"))
    if resolved_index:
        index_name = resolved_index
    records = _prepare_records(
        document_id,
        filename,
        chunks,
        workflow_id=workflow_id,
        instance=(doc or {}).get("instance"),
    )
    payload_path = _write_json_temp(records)
    try:
        payload_uri, payload_size, payload_mime = _upload_file_to_minio(
            payload_path, workflow_id, "marqo_payload_export", "marqo_payload.json",
            instance=(doc or {}).get("instance"),
        )
    finally:
        if os.path.exists(payload_path):
            os.remove(payload_path)
    latest_job = db.get_latest_document_job(workflow_id)
    db.add_document_artifact(
        workflow_id=workflow_id,
        job_id=latest_job["id"] if latest_job else None,
        artifact_type="marqo_payload_export",
        stage="ingesting",
        storage_uri=payload_uri,
        mime_type=payload_mime,
        filename="marqo_payload.json",
        size_bytes=payload_size,
        metadata={"record_count": len(records), "index_name": index_name},
    )
    result = await ingest_to_marqo(records, marqo_url=marqo_url, index_name=index_name, batch_size=batch_size)
    db.upsert_document_index_status(
        workflow_id=workflow_id,
        index_name=index_name,
        marqo_doc_id=document_id,
        chunk_count_indexed=result.get("records_ingested", 0),
        last_indexed_at=datetime.utcnow().isoformat(),
        last_verified_at=datetime.utcnow().isoformat(),
        schema_version="passage-v1",
        status="indexed",
        details=result.get("index_stats"),
    )
    return result


@activity.defn
async def update_document_state(
    workflow_id: str,
    stage: str,
    page_count: int = 0,
    chunk_count: int = 0,
    error_message: str = None,
) -> dict:
    """Update document state in SQLite."""
    from . import db

    activity.logger.info(f"Updating state for {workflow_id}: stage={stage}")
    db.update_document_stage(
        workflow_id=workflow_id,
        stage=stage,
        page_count=page_count,
        chunk_count=chunk_count,
        error_message=error_message,
    )
    latest_job = db.get_latest_document_job(workflow_id)
    if latest_job:
        job_updates = {"current_stage": stage}
        if stage in {"ocr_review", "translation_review", "chunk_review", "ready_for_ingestion"}:
            job_updates["status"] = "waiting_review"
        elif stage == "completed":
            job_updates["status"] = "completed"
            job_updates["completed_at"] = datetime.utcnow().isoformat()
        elif stage == "failed":
            job_updates["status"] = "failed"
            job_updates["completed_at"] = datetime.utcnow().isoformat()
            job_updates["error_message"] = error_message
        else:
            job_updates["status"] = "running"
        db.update_document_job(latest_job["id"], **job_updates)
    return {"updated": True, "stage": stage}


@activity.defn
async def persist_document_content(workflow_id: str, pages: list[dict], chunks: list[dict]) -> dict:
    """Persist pages and chunks to SQLite."""
    from . import db

    activity.logger.info(f"Persisting content for {workflow_id}: {len(pages)} pages, {len(chunks)} chunks")
    db.save_pages(workflow_id, pages)
    db.save_chunks(workflow_id, chunks)
    return {"persisted": True, "pages": len(pages), "chunks": len(chunks)}


@activity.defn
async def auto_tag_chunks_from_db(workflow_id: str, filename: str = "") -> dict:
    """Auto-assign domain tags to chunks using the configured LLM tagger."""
    from . import db
    from .domain_tags.base import validate_tags_against_taxonomy
    from .domain_tags.gemma_tagger import auto_tag_chunks
    from .domain_tags.service import get_domain_tagger, load_domain_tagging_config

    config = load_domain_tagging_config()
    if not config.enabled:
        activity.logger.info("Domain tagging disabled; skipping workflow %s", workflow_id)
        return {"tagged_chunks": 0, "skipped": True}

    chunks = db.get_chunks(workflow_id, include_excluded=True)
    if not chunks:
        return {"tagged_chunks": 0, "skipped": True}

    doc = db.get_document(workflow_id) or {}
    doc_context_parts = [
        doc.get("source_manifest_name") or "",
        doc.get("display_name") or "",
    ]
    doc_context = " | ".join(part for part in doc_context_parts if part)

    tagger = get_domain_tagger(config)
    tagged_map = await auto_tag_chunks(
        chunks,
        filename=filename or doc.get("filename") or "",
        doc_context=doc_context,
        tagger=tagger,
        log=activity.logger.info,
    )

    db.delete_auto_chunk_tags(workflow_id)
    tagged_chunks = 0
    total_tags = 0
    for chunk_num, tags in tagged_map.items():
        if config.strict_taxonomy:
            tags = validate_tags_against_taxonomy(tags, strict=True)
        if not tags:
            continue
        db.replace_chunk_tags(
            workflow_id,
            chunk_num,
            [{"dimension": t.dimension, "value": t.value} for t in tags],
            source="auto",
        )
        tagged_chunks += 1
        total_tags += len(tags)

    activity.logger.info(
        "Auto domain tagging complete for %s: %s chunks, %s tags",
        workflow_id,
        tagged_chunks,
        total_tags,
    )
    return {"tagged_chunks": tagged_chunks, "total_tags": total_tags, "skipped": False}


@activity.defn
async def detect_and_translate_pages(
    pages: list[dict],
    target_language: str = "en",
    source_language: str = None,
) -> list[dict]:
    """Detect language and translate non-English pages."""
    return await _detect_and_translate_impl(pages, target_language=target_language, source_language=source_language)


@activity.defn
async def detect_and_translate_pages_from_db(
    workflow_id: str,
    target_language: str = "en",
    source_language: str = None,
) -> dict:
    """Detect and translate pages loaded from SQLite; persist updated pages back to SQLite."""
    from . import db

    pages = db.get_pages(workflow_id)
    translated = await _detect_and_translate_impl(pages, target_language=target_language, source_language=source_language)
    db.save_pages(workflow_id, translated)
    translated_count = sum(1 for p in translated if p.get("translated_markdown"))
    latest_job = db.get_latest_document_job(workflow_id)
    translation_config = load_translation_config(target_language=target_language)
    translation_instance = (db.get_document(workflow_id) or {}).get("instance")
    translated_json_path = _write_json_temp(translated)
    try:
        translated_uri, translated_size, translated_mime = _upload_file_to_minio(
            translated_json_path, workflow_id, "translation_pages_json", "translated_pages.json",
            instance=translation_instance,
        )
    finally:
        if os.path.exists(translated_json_path):
            os.remove(translated_json_path)
    db.add_document_artifact(
        workflow_id=workflow_id,
        job_id=latest_job["id"] if latest_job else None,
        artifact_type="translation_pages_json",
        stage="translation_review",
        storage_uri=translated_uri,
        mime_type=translated_mime,
        filename="translated_pages.json",
        size_bytes=translated_size,
        metadata={
            "page_count": len(translated),
            "translated_count": translated_count,
            "translation_provider": translation_config.provider,
            "translation_model": translation_config.model,
            "translation_target_language": target_language,
            "translation_run_id": str(uuid4()),
        },
    )
    return {"page_count": len(translated), "translated_count": translated_count}
